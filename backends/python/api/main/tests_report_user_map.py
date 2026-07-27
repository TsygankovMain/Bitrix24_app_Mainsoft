"""_get_user_map строит карту имён из локальной БД (portal_user), а не из
Bitrix user.get (Фаза 2 sync-offload: убирает 3-7с "user_map" на отчётах)."""
import io
import json
from unittest.mock import MagicMock, PropertyMock, patch

import openpyxl
from django.test import Client, RequestFactory, TestCase, override_settings

from . import views
from .models import Bitrix24Account, Portal, PortalUser


class GetUserMapReadsFromPortalUserTest(TestCase):
    def setUp(self):
        self.account = Bitrix24Account.objects.create(
            b24_user_id=1, is_b24_user_admin=True, member_id="m-map-1",
            is_master_account=True, domain_url="example.bitrix24.ru",
            status="active", application_version=1,
        )

    def _request(self):
        request = RequestFactory().get("/api/report-employee-project")
        request.bitrix24_account = self.account
        return request

    def test_builds_map_from_local_db_without_bitrix_call(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="1", name="Иван", last_name="Петров", active=True)
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="2", name="Анна", last_name="Сидорова", active=False)

        result = views._get_user_map(self._request(), {"1", "2"})

        self.assertEqual(result, {"1": "Петров Иван", "2": "Сидорова Анна"})

    def test_missing_user_id_is_simply_absent_from_map(self):
        result = views._get_user_map(self._request(), {"999"})
        self.assertEqual(result, {})  # resolve_employee_name падает на fallback "Сотрудник 999"

    def test_empty_user_ids_returns_empty_dict(self):
        self.assertEqual(views._get_user_map(self._request(), set()), {})

    def test_scoped_by_tenant_other_account_users_not_leaked(self):
        other = Bitrix24Account.objects.create(
            b24_user_id=2, is_b24_user_admin=True, member_id="m-map-2",
            is_master_account=True, domain_url="other.bitrix24.ru",
            status="active", application_version=1,
        )
        PortalUser.objects.create(bitrix24_account=other, bitrix_id="1", name="Чужой", last_name="Юзер", active=True)

        result = views._get_user_map(self._request(), {"1"})
        self.assertEqual(result, {})


class GetUserMapNormalizesNonCanonicalIdsTest(TestCase):
    """Ревью Задачи 4 (эскалировано до обязательного фикса): историчные
    TimesheetItem.employee_id могут быть неканоничными ("[12]", "12.0") —
    старый fetch_users резолвил их через numeric_to_aliases, поэтому
    _get_user_map обязан нормализовать id ПЕРЕД запросом к PortalUser (см.
    extract_bitrix_user_id — тот же конвертер, которым UserSyncService
    пишет PortalUser.bitrix_id) и отдавать канонический ключ (его же ищет
    resolve_employee_name первым делом через normalize_employee_id)."""

    def setUp(self):
        self.account = Bitrix24Account.objects.create(
            b24_user_id=1, is_b24_user_admin=True, member_id="m-map-3",
            is_master_account=True, domain_url="example-norm.bitrix24.ru",
            status="active", application_version=1,
        )

    def _request(self):
        request = RequestFactory().get("/api/report-employee-project")
        request.bitrix24_account = self.account
        return request

    def test_bracket_form_employee_id_resolves_to_canonical_key(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="12", name="Игорь", last_name="Смирнов", active=True)

        result = views._get_user_map(self._request(), {"[12]"})

        self.assertEqual(result, {"12": "Смирнов Игорь"})

    def test_float_form_employee_id_resolves_to_canonical_key(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="12", name="Игорь", last_name="Смирнов", active=True)

        result = views._get_user_map(self._request(), {"12.0"})

        self.assertEqual(result, {"12": "Смирнов Игорь"})

    def test_canonical_employee_id_still_resolves_name(self):
        """Регресс: нормализация не должна ломать уже-канонический путь."""
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="12", name="Игорь", last_name="Смирнов", active=True)

        result = views._get_user_map(self._request(), {"12"})

        self.assertEqual(result, {"12": "Смирнов Игорь"})

    def test_both_name_fields_empty_falls_back_to_bitrix_id(self):
        """Minor из ревью: пустые name/last_name -> значение падает на bitrix_id."""
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="5", name="", last_name="", active=True)

        result = views._get_user_map(self._request(), {"5"})

        self.assertEqual(result, {"5": "5"})


# ---------------------------------------------------------------------------
# export_raw_data — единственный эндпоинт, оставшийся в Фазе 2 со своим
# ручным батчингом user.get (отложенный финдинг ревью Задачи 4). После
# рефакторинга он резолвит имена тем же _get_user_map, что и 14 report_*:
# локальный PortalUser + scope_to_tenant + extract_bitrix_user_id, но со
# своим фолбэком «нет имени -> сырой id» (а не «Сотрудник <id>»).
# ---------------------------------------------------------------------------

EMPLOYEE_FIELD = "ufCrm5Employee"
HOURS_FIELD = "ufCrm5Hours"

SP_FIELDS_META = [
    {"id": EMPLOYEE_FIELD, "title": "Сотрудник", "type": "employee"},
    {"id": HOURS_FIELD, "title": "Часы", "type": "double"},
]


class ExportRawDataMixin:
    """Прогон export_raw_data через HTTP с замоканным Bitrix."""

    def _call_method(self, method, params=None):
        """Заглушка Bitrix: crm.item.list отдаёт items, user.get — пустой
        result (как будто Bitrix имён не вернул). Все вызовы пишутся в
        self.calls, чтобы проверить отсутствие user.get."""
        self.calls.append(method)
        if method == "crm.item.list":
            return {"result": {"items": list(self.items)}, "total": len(self.items)}
        return {"result": []}

    def _export(self, items, fields=None):
        """POST /api/export-raw-data -> лист openpyxl."""
        self.items = items

        bitrix_client = MagicMock()
        bitrix_client._bitrix_token.call_method.side_effect = self._call_method

        config_service = MagicMock()
        config_service.get_configuration_sync.return_value = {
            "sp_entity_type_id": 1032,
            "fields_mapping": {"data": "ufCrm5Date"},
        }
        config_service.get_sp_fields_sync.return_value = SP_FIELDS_META

        body = {"date_from": "", "date_to": "", "date_type": "reflection",
                "fields": fields if fields is not None else [EMPLOYEE_FIELD, HOURS_FIELD]}

        with patch.object(Bitrix24Account, "client", new_callable=PropertyMock, return_value=bitrix_client), \
                patch.object(views, "ConfigurationService", return_value=config_service):
            response = self.client.post(
                "/api/export-raw-data",
                data=json.dumps(body),
                content_type="application/json",
                HTTP_AUTHORIZATION=f"Bearer {self.account.create_jwt_token()}",
            )

        self.assertEqual(response.status_code, 200, response.content[:500])
        return openpyxl.load_workbook(io.BytesIO(response.content)).active


class ExportRawDataUsesLocalUserMapTest(ExportRawDataMixin, TestCase):
    """export_raw_data берёт имена из PortalUser через _get_user_map и не
    ходит в Bitrix user.get."""

    def setUp(self):
        self.client = Client()
        self.account = Bitrix24Account.objects.create(
            b24_user_id=1, is_b24_user_admin=True, member_id="m-export-1",
            is_master_account=True, domain_url="export.bitrix24.ru",
            status="active", application_version=1,
        )
        self.calls = []

    def test_employee_name_comes_from_local_portal_user(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="1",
                                  name="Иван", last_name="Петров", active=True)

        ws = self._export([{EMPLOYEE_FIELD: "1", HOURS_FIELD: "8"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "Петров Иван")

    def test_does_not_call_bitrix_user_get(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="1",
                                  name="Иван", last_name="Петров", active=True)

        self._export([{EMPLOYEE_FIELD: "1", HOURS_FIELD: "8"}])

        self.assertNotIn("user.get", self.calls)

    def test_inactive_employee_name_still_resolves(self):
        """Историчные строки по уволенным: PortalUser хранит и неактивных."""
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="7",
                                  name="Анна", last_name="Сидорова", active=False)

        ws = self._export([{EMPLOYEE_FIELD: "7", HOURS_FIELD: "4"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "Сидорова Анна")

    def test_multiple_employee_values_are_joined(self):
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="1",
                                  name="Иван", last_name="Петров", active=True)
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="2",
                                  name="Анна", last_name="Сидорова", active=True)

        ws = self._export([{EMPLOYEE_FIELD: [1, 2], HOURS_FIELD: "8"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "Петров Иван; Сидорова Анна")

    def test_non_canonical_employee_value_resolves_to_name(self):
        """_get_user_map отдаёт КАНОНИЧНЫЕ ключи (extract_bitrix_user_id), поэтому
        поиск по сырому str(val) для неканоничных значений ("12.0") промахнётся —
        резолв обязан искать сначала по нормализованному id."""
        PortalUser.objects.create(bitrix24_account=self.account, bitrix_id="12",
                                  name="Игорь", last_name="Смирнов", active=True)

        ws = self._export([{EMPLOYEE_FIELD: 12.0, HOURS_FIELD: "8"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "Смирнов Игорь")

    def test_unknown_employee_id_falls_back_to_raw_id(self):
        """Фолбэк export_raw_data — сырой id, а НЕ "Сотрудник 999" из
        resolve_employee_name: _get_user_map ключ не вернёт вовсе."""
        ws = self._export([{EMPLOYEE_FIELD: "999", HOURS_FIELD: "8"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "999")

    def test_empty_employee_value_stays_empty(self):
        ws = self._export([{EMPLOYEE_FIELD: None, HOURS_FIELD: "8"}])

        self.assertIsNone(ws.cell(row=2, column=1).value)

    def test_other_account_user_name_is_not_leaked(self):
        """Мультитенантность: имя из чужого аккаунта не должно попасть в выгрузку."""
        other = Bitrix24Account.objects.create(
            b24_user_id=2, is_b24_user_admin=True, member_id="m-export-2",
            is_master_account=True, domain_url="other-export.bitrix24.ru",
            status="active", application_version=1,
        )
        PortalUser.objects.create(bitrix24_account=other, bitrix_id="1",
                                  name="Чужой", last_name="Юзер", active=True)

        ws = self._export([{EMPLOYEE_FIELD: "1", HOURS_FIELD: "8"}])

        self.assertEqual(ws.cell(row=2, column=1).value, "1")


@override_settings(USE_PORTAL_SCOPING=True)
class ExportRawDataPortalScopingTest(ExportRawDataMixin, TestCase):
    """При включённом portal-скоупинге выгрузка видит справочник своей компании
    и не видит чужой (scope_to_tenant внутри _get_user_map)."""

    def setUp(self):
        self.client = Client()
        self.calls = []
        self.portal = Portal.objects.create(member_id="m-export-portal-1", domain_url="p1.bitrix24.ru")
        self.other_portal = Portal.objects.create(member_id="m-export-portal-2", domain_url="p2.bitrix24.ru")
        self.account = Bitrix24Account.objects.create(
            b24_user_id=1, is_b24_user_admin=True, member_id="m-export-portal-1",
            is_master_account=True, domain_url="p1.bitrix24.ru",
            status="active", application_version=1, portal=self.portal,
        )
        self.other_account = Bitrix24Account.objects.create(
            b24_user_id=2, is_b24_user_admin=True, member_id="m-export-portal-2",
            is_master_account=True, domain_url="p2.bitrix24.ru",
            status="active", application_version=1, portal=self.other_portal,
        )

    def test_own_portal_resolves_and_other_portal_does_not_leak(self):
        PortalUser.objects.create(bitrix24_account=self.account, portal=self.portal,
                                  bitrix_id="1", name="Иван", last_name="Петров", active=True)
        PortalUser.objects.create(bitrix24_account=self.other_account, portal=self.other_portal,
                                  bitrix_id="2", name="Чужой", last_name="Юзер", active=True)

        ws = self._export([
            {EMPLOYEE_FIELD: "1", HOURS_FIELD: "8"},
            {EMPLOYEE_FIELD: "2", HOURS_FIELD: "4"},
        ])

        self.assertEqual(ws.cell(row=2, column=1).value, "Петров Иван")
        self.assertEqual(ws.cell(row=3, column=1).value, "2")

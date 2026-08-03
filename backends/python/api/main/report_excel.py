"""
Генерация структурированного Excel-файла для отчёта «Учет по проектам/задачам».

Сохраняет иерархию (проект → задача → подзадача → сотрудник → метка времени):
- объединённая шапка с названием отчёта и параметрами периода/фильтров;
- цветовая заливка и отступ по уровням;
- сворачивание групп (outline-группировка строк);
- числа как настоящие числа (формат «0.0») — суммируются в Excel;
- финальная строка ИТОГО.

Структура входных данных `nodes` совпадает с тем, что отдаёт
ReportService().generate_project_task_employees(...):
узел = {name, total_hours, billable_hours, non_billable_hours,
        children?: [...], employees?: [{name, ..., items?: [...]}]}.
"""

import io
from typing import Any, Dict, List, Optional, Sequence

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Колонки: A — название, B — всего, C — учтено, D — не учтено
_HOURS_FORMAT = "0.0"
_THIN = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Заливки уровней (HEX без #)
_FILL_TITLE = PatternFill("solid", fgColor="1F2937")
_FILL_SUBTITLE = PatternFill("solid", fgColor="374151")
_FILL_HEAD = PatternFill("solid", fgColor="E5E7EB")
_FILL_PROJECT = PatternFill("solid", fgColor="ECFCCB")
_FILL_TASK = PatternFill("solid", fgColor="F1F5F9")
_FILL_SUBTASK = PatternFill("solid", fgColor="F8FAFC")
_FILL_EMPLOYEE = PatternFill("solid", fgColor="FFFFFF")
_FILL_ITEM = PatternFill("solid", fgColor="FBFDFF")
_FILL_TOTAL = PatternFill("solid", fgColor="CBD5E1")

_COLOR_BILL = "047857"
_COLOR_NONBILL = "BE123C"

# ---------------------------------------------------------------------------
# Лимит строк (мягкий предохранитель объёма выгрузки)
# ---------------------------------------------------------------------------

MAX_EXPORT_ROWS = 50000


class ExportTooLargeError(Exception):
    """Выгрузка превышает мягкий лимит строк. View ловит и отдаёт HTTP 400."""

    def __init__(self, rows: int, limit: int = MAX_EXPORT_ROWS):
        self.rows = rows
        self.limit = limit
        super().__init__(
            f"Слишком большой период или выборка для выгрузки "
            f"(строк: {rows} > {limit}). Сузьте период или фильтры."
        )


def _count_hierarchy_rows(roots) -> int:
    """Считает строки данных иерархии (узлы + листовые items + employees+их items)."""
    total = 0

    def walk(node):
        nonlocal total
        total += 1
        for ch in node.get("children") or []:
            walk(ch)
        for emp in node.get("employees") or []:
            total += 1
            total += len(emp.get("items") or [])
        total += len(node.get("items") or [])

    for r in roots:
        walk(r)
    return total


def _num(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


# Символы, с которых начинаются Excel-формулы/инъекции
_FORMULA_STARTERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell_text(value: Any) -> Any:
    """Нейтрализует Excel-инъекцию формул: строки, начинающиеся с опасных символов,
    получают префикс апостроф ('), который Excel воспринимает как «текст, не формула».
    Нечисловые значения и None возвращаются без изменений."""
    if isinstance(value, str) and value.startswith(_FORMULA_STARTERS):
        return "'" + value
    return value


def _write_row(
    ws: Worksheet,
    row: int,
    *,
    name: str,
    total: float,
    billable: float,
    non_billable: float,
    depth: int,
    fill: PatternFill,
    bold: bool = False,
    italic: bool = False,
    name_color: Optional[str] = None,
    outline_level: int = 0,
) -> None:
    """Записывает одну строку из 4 колонок с форматированием и уровнем группировки."""
    name_cell = ws.cell(row=row, column=1, value=_safe_cell_text(name))
    name_cell.alignment = Alignment(horizontal="left", vertical="center", indent=depth)
    name_cell.font = Font(bold=bold, italic=italic, color=name_color or "0F172A")
    name_cell.fill = fill
    name_cell.border = _BORDER

    for col, value, color in (
        (2, total, "0F172A"),
        (3, billable, _COLOR_BILL),
        (4, non_billable, _COLOR_NONBILL),
    ):
        cell = ws.cell(row=row, column=col, value=round(_num(value), 2))
        cell.number_format = _HOURS_FORMAT
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.font = Font(bold=bold, color=color)
        cell.fill = fill
        cell.border = _BORDER

    if outline_level > 0:
        ws.row_dimensions[row].outline_level = min(outline_level, 7)


def _format_iso_date(value: Any) -> str:
    """ISO-дата (YYYY-MM-DD...) -> ДД.ММ.ГГГГ, без зависимости от локали сервера."""
    if not isinstance(value, str) or len(value) < 10:
        return ""
    y, m, d = value[0:4], value[5:7], value[8:10]
    if value[4] == "-" and value[7] == "-" and y.isdigit() and m.isdigit() and d.isdigit():
        return f"{d}.{m}.{y}"
    return value[:10]


def _write_item(ws: Worksheet, row: int, item: Dict[str, Any], depth: int) -> None:
    # Строка списания в выгрузке показывает ОПИСАНИЕ, а не название задачи —
    # тот же порядок, что и на экране (ProjectTaskReportEmployeeRow.vue).
    # Название задачи повторяется у каждого списания группы и прячет то
    # единственное, чем они отличаются. Порядок был обратным до 29.07.2026.
    name = item.get("opisanie") or item.get("nazvanie_zadachi") or "Без описания"
    formatted_date = _format_iso_date(item.get("data"))
    if formatted_date:
        name = f"{name} · {formatted_date}"
    hours = _num(item.get("kolichestvo_chasov"))
    is_billable = bool(item.get("uchitivaem"))
    _write_row(
        ws,
        row,
        name=name,
        total=hours,
        billable=hours if is_billable else 0.0,
        non_billable=0.0 if is_billable else hours,
        depth=depth,
        fill=_FILL_ITEM,
        italic=True,
        name_color="64748B",
        outline_level=depth,
    )


def _write_employee(ws: Worksheet, start_row: int, employee: Dict[str, Any], depth: int) -> int:
    row = start_row
    _write_row(
        ws,
        row,
        name=employee.get("name") or "—",
        total=employee.get("total_hours"),
        billable=employee.get("billable_hours"),
        non_billable=employee.get("non_billable_hours"),
        depth=depth,
        fill=_FILL_EMPLOYEE,
        name_color="334155",
        outline_level=depth,
    )
    row += 1
    for item in employee.get("items") or []:
        _write_item(ws, row, item, depth + 1)
        row += 1
    return row


def _write_node(ws: Worksheet, start_row: int, node: Dict[str, Any], depth: int) -> int:
    """Рекурсивно пишет узел (проект/задача/подзадача) и его потомков. Возвращает следующий свободный row."""
    is_project = depth == 0
    if is_project:
        fill, bold, name_color = _FILL_PROJECT, True, "3F6212"
    elif depth == 1:
        fill, bold, name_color = _FILL_TASK, True, "1E293B"
    else:
        fill, bold, name_color = _FILL_SUBTASK, False, "334155"

    row = start_row
    _write_row(
        ws,
        row,
        name=node.get("name") or "—",
        total=node.get("total_hours"),
        billable=node.get("billable_hours"),
        non_billable=node.get("non_billable_hours"),
        depth=depth,
        fill=fill,
        bold=bold,
        name_color=name_color,
        outline_level=depth,
    )
    row += 1

    for child in node.get("children") or []:
        row = _write_node(ws, row, child, depth + 1)
    for employee in node.get("employees") or []:
        row = _write_employee(ws, row, employee, depth + 1)
    return row


def build_project_task_workbook(
    nodes: Sequence[Dict[str, Any]],
    *,
    date_from: str = "",
    date_to: str = "",
    filters_label: str = "",
) -> io.BytesIO:
    """Строит xlsx-файл отчёта и возвращает BytesIO (указатель в начале).

    Остаётся в обычном режиме (write_only несовместим с outline-группировкой строк
    через ws.row_dimensions[row].outline_level). Только guard на объём.
    """
    n_rows = _count_hierarchy_rows(list(nodes))
    if n_rows > MAX_EXPORT_ROWS:
        raise ExportTooLargeError(n_rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проект-Задача"

    # Группировка: кнопка сворачивания на строке-родителе (она ВЫШЕ потомков)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # --- Шапка отчёта (объединённые ячейки) ---
    period = f"{date_from} — {date_to}".strip(" —")
    title = "Учет по проектам/задачам"
    if period:
        title = f"{title} · период {period}"

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=_safe_cell_text(title))
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = _FILL_TITLE
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:D2")
    subtitle = filters_label or "Сотрудники: все · Проекты: все"
    c2 = ws.cell(row=2, column=1, value=_safe_cell_text(subtitle))
    c2.font = Font(color="CBD5E1", size=10)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2.fill = _FILL_SUBTITLE

    # --- Шапка колонок ---
    headers = ["Название", "Всего, ч", "Учтено, ч", "Не учтено, ч"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(bold=True, color="111827")
        cell.fill = _FILL_HEAD
        cell.border = _BORDER
        cell.alignment = Alignment(
            horizontal="left" if col == 1 else "right", vertical="center"
        )

    # --- Данные ---
    row = 4
    grand_total = grand_bill = grand_nonbill = 0.0
    for node in nodes:
        grand_total += _num(node.get("total_hours"))
        grand_bill += _num(node.get("billable_hours"))
        grand_nonbill += _num(node.get("non_billable_hours"))
        row = _write_node(ws, row, node, 0)

    # --- ИТОГО ---
    _write_row(
        ws,
        row,
        name="ИТОГО",
        total=grand_total,
        billable=grand_bill,
        non_billable=grand_nonbill,
        depth=0,
        fill=_FILL_TOTAL,
        bold=True,
    )

    # Ширина колонок и закрепление заголовков
    widths = {1: 55, 2: 12, 3: 12, 4: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_hierarchy_workbook(roots, *, title, date_from="", date_to="",
                             value_columns=(("Всего, ч", "total_hours"),
                                            ("Учтено, ч", "billable_hours"),
                                            ("Не учтено, ч", "non_billable_hours"))):
    """Иерархия с outline-группировкой строк. Остаётся в обычном режиме (write_only
    несовместим с ws.row_dimensions[row].outline_level). Только guard на объём."""
    n_rows = _count_hierarchy_rows(roots)
    if n_rows > MAX_EXPORT_ROWS:
        raise ExportTooLargeError(n_rows)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Отчёт"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ncols = 1 + len(value_columns)
    last_col = get_column_letter(ncols)
    period = f"{date_from} — {date_to}".strip(" —")
    full_title = f"{title} · период {period}" if period else title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(1, 1, full_title); c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = _FILL_TITLE; c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    h = ws.cell(2, 1, "Название"); h.font = Font(bold=True, color="111827"); h.fill = _FILL_HEAD
    h.border = _BORDER; h.alignment = Alignment(horizontal="left", vertical="center")
    for i, (label, _) in enumerate(value_columns):
        cell = ws.cell(2, 2 + i, label); cell.font = Font(bold=True, color="111827")
        cell.fill = _FILL_HEAD; cell.border = _BORDER; cell.alignment = Alignment(horizontal="right", vertical="center")
    row = 3
    totals = [0.0] * len(value_columns)

    def _write(node, depth):
        nonlocal row
        fill = _FILL_PROJECT if depth == 0 else (_FILL_TASK if depth == 1 else _FILL_SUBTASK)
        bold = depth <= 1
        nc = ws.cell(row, 1, _safe_cell_text(node.get("name") or "—"))
        nc.alignment = Alignment(horizontal="left", vertical="center", indent=depth)
        nc.font = Font(bold=bold); nc.fill = fill; nc.border = _BORDER
        for i, (_, key) in enumerate(value_columns):
            cell = ws.cell(row, 2 + i, round(_num(node.get(key)), 2))
            cell.number_format = _HOURS_FORMAT; cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.font = Font(bold=bold); cell.fill = fill; cell.border = _BORDER
        if depth > 0:
            ws.row_dimensions[row].outline_level = min(depth, 7)
        row += 1
        for ch in node.get("children") or []:
            _write(ch, depth + 1)
        # Листовые записи времени (как на экране) — только для дефолтного hours-layout
        if len(value_columns) == 3:
            for item in node.get("items") or []:
                _write_item(ws, row, item, depth + 1)
                row += 1

    for node in roots:
        for i, (_, key) in enumerate(value_columns):
            totals[i] += _num(node.get(key))
        _write(node, 0)

    tc = ws.cell(row, 1, "ИТОГО"); tc.font = Font(bold=True); tc.fill = _FILL_TOTAL; tc.border = _BORDER
    for i, t in enumerate(totals):
        cell = ws.cell(row, 2 + i, round(t, 2)); cell.number_format = _HOURS_FORMAT
        cell.font = Font(bold=True); cell.fill = _FILL_TOTAL
        cell.alignment = Alignment(horizontal="right", vertical="center"); cell.border = _BORDER
    ws.column_dimensions["A"].width = 55
    for idx in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.freeze_panes = "A3"
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output


def build_matrix_workbook(header_days, rows, *, title, date_from="", date_to=""):
    """Матрица сотрудник×день. write_only режим (линейный расход памяти)."""
    if len(rows) > MAX_EXPORT_ROWS:
        raise ExportTooLargeError(len(rows))

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Нагрузка")
    days = [(d.get("date") if isinstance(d, dict) else d) for d in header_days]
    ncols = 1 + len(days) + 1
    # freeze и ширины ставим ДО append (требование write_only)
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 24
    for idx in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 10

    period = f"{date_from} — {date_to}".strip(" —")
    full_title = f"{title} · период {period}" if period else title

    def _styled(value, *, number=False, bold=False, fill=None, align="right"):
        cell = WriteOnlyCell(ws, value=value)
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        if number:
            cell.number_format = _HOURS_FORMAT
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = _BORDER
        return cell

    # Строка 1: заголовок (без merge — write_only merge ненадёжен; пишем в A1)
    title_cell = WriteOnlyCell(ws, value=_safe_cell_text(full_title))
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = _FILL_TITLE
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.append([title_cell])

    # Строка 2: шапка
    head = [_styled("Сотрудник", bold=True, fill=_FILL_HEAD, align="left")]
    for day in days:
        head.append(_styled(_format_iso_date(day) or str(day), bold=True, fill=_FILL_HEAD))
    head.append(_styled("Итого", bold=True, fill=_FILL_HEAD))
    ws.append(head)

    # Данные
    col_tot = [0.0] * len(days)
    grand = 0.0
    for r in rows:
        name = (r.get("employee") or {}).get("name") or "—"
        row_cells = [_styled(_safe_cell_text(name), align="left")]
        rowsum = 0.0
        cells = r.get("days") or {}
        for i, day in enumerate(days):
            cd = cells.get(day) or {}
            v = _num(cd.get("total")) if isinstance(cd, dict) else _num(cd)
            row_cells.append(_styled(round(v, 2) if v else None, number=True))
            rowsum += v
            col_tot[i] += v
        row_cells.append(_styled(round(rowsum, 2), number=True, bold=True, fill=_FILL_TOTAL))
        grand += rowsum
        ws.append(row_cells)

    # ИТОГО
    total_cells = [_styled("ИТОГО", bold=True, fill=_FILL_TOTAL, align="left")]
    for ct in col_tot:
        total_cells.append(_styled(round(ct, 2), number=True, bold=True, fill=_FILL_TOTAL))
    total_cells.append(_styled(round(grand, 2), number=True, bold=True, fill=_FILL_TOTAL))
    ws.append(total_cells)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


_TABLE_FMT = {"text": "@", "hours": "0.0", "money": "#,##0", "percent": "0.0%", "int": "0"}


def build_table_workbook(columns, rows, *, title, date_from="", date_to="", total_row=None):
    """Плоская таблица. write_only режим (линейный расход памяти)."""
    if len(rows) > MAX_EXPORT_ROWS:
        raise ExportTooLargeError(len(rows))

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Отчёт")
    ncols = len(columns)
    # freeze и ширины ставим ДО append (требование write_only)
    ws.freeze_panes = "A3"
    for i, col in enumerate(columns):
        ws.column_dimensions[get_column_letter(1 + i)].width = col.get("width", 18)

    period = f"{date_from} — {date_to}".strip(" —")
    full_title = f"{title} · период {period}" if period else title

    def _cell(value, *, number_fmt=None, bold=False, fill=None, align="left"):
        cell = WriteOnlyCell(ws, value=value)
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        if number_fmt:
            cell.number_format = number_fmt
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = _BORDER
        return cell

    # Строка 1: заголовок (без merge — write_only merge ненадёжен; пишем в A1)
    title_cell = WriteOnlyCell(ws, value=_safe_cell_text(full_title))
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = _FILL_TITLE
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.append([title_cell])

    # Строка 2: шапка
    header = []
    for col in columns:
        align = "left" if col.get("fmt", "text") == "text" else "right"
        header.append(_cell(col["label"], bold=True, fill=_FILL_HEAD, align=align))
    ws.append(header)

    def _row_cells(r, *, bold=False, fill=None):
        out_cells = []
        for col in columns:
            fmt = col.get("fmt", "text")
            val = r.get(col["key"])
            if fmt == "text" or val is None:
                out_cells.append(_cell(
                    "" if val is None else _safe_cell_text(str(val)),
                    bold=bold, fill=fill, align="left",
                ))
            else:
                out_cells.append(_cell(
                    _num(val), number_fmt=_TABLE_FMT[fmt],
                    bold=bold, fill=fill, align="right",
                ))
        return out_cells

    for r in rows:
        ws.append(_row_cells(r))
    if total_row:
        ws.append(_row_cells(total_row, bold=True, fill=_FILL_TOTAL))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

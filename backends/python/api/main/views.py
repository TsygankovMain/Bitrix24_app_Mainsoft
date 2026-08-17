from collections import defaultdict
from datetime import timedelta

from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.clickjacking import xframe_options_exempt

from .utils.decorators import auth_required, log_errors, rate_limit
from .utils.decorators.sync_lock import sync_lock, account_sync_lock, SyncLockBusy
from .utils import AuthorizedRequest
from .models import ApplicationInstallation, TimesheetItem, RequestLog, SystemLog, ProjectCard, PortalUser

import logging
import json
import openpyxl
from openpyxl.styles import Font, Alignment

from config import load_config
from .services import (
    BitrixDataService,
    ReportService,
    TimesheetSyncService,
    ConfigurationService,
    ProjectCardService,
    ProjectSyncService,
    ProjectStageAutomationService,
    invalidate_project_runtime_caches,
)
from .installation_service import InstallationService, InstallationError
from .timesheet_sync_service import resolve_sync_mode
from .tenant_scoping import scope_to_tenant
from .employee_ids import extract_bitrix_user_id
from .perf import ReportProfiler
from .report_queries import (
    TREE_REPORT_FIELDS,
    build_filtered_timesheet_queryset,
    build_project_filter_options,
    build_project_title_lookups,
    build_tree_report_items,
    materialize_rows,
    resolve_project_name_for_row,
)
from .report_excel import (
    build_project_task_workbook,
    build_hierarchy_workbook,
    build_matrix_workbook,
    build_table_workbook,
    ExportTooLargeError,
    _safe_cell_text,
)
from .inn_backfill_service import InnBackfillService
from .company_search_service import CompanySearchService
from .project_creation_service import ProjectCreationService

__all__ = [
    "root",
    "health",
    "health_check",
    "serve_spa",
    "get_enum",
    "get_list",
    "install",
    "get_token",
    "get_filter_options",
    "get_filter_employees",
    "get_filter_projects",
    "get_support_status",
    "connect_support_line",
    "get_project_board",
    "get_project_board_meta",
    "get_project_board_card",
    "sync_project_board",
    "create_project_board",
    "update_project_board",
    "update_project_board_stage",
    "archive_project_board",
    "run_project_board_daily_check",
    "get_project_board_companies",
    "search_project_board_companies",
    "list_my_companies",
    "get_homepage_portfolio",
    "get_internal_lists",
    "report_employee_project",
    "report_project_employee",
    "report_daily_workload",
    "report_project_task_employee",
    "report_project_task_employee_export",
    "report_employee_project_export",
    "report_project_employee_export",
    "report_daily_workload_export",
    "report_revenue_leakage_export",
    "report_time_entry_discipline_export",
    "report_focus_analysis_export",
    "inn_backfill_scan",
    "inn_backfill_apply",
    "inn_backfill_project_items",
    "projects_health",
    "report_revenue_leakage",
    "report_time_entry_discipline",
    "report_focus_analysis",
    "timesheet_sync",
    "timesheet_sync_status",
    "timesheet_list",
    "get_users",
    "get_configuration",
    "save_configuration",
    "get_smart_processes",
    "get_sp_fields",
    "get_project_spa_validation",
    "get_project_spa_stages",
    "run_project_spa_backfill",
    "get_request_logs",
    "get_system_logs",
    "create_smart_process",
    "create_fields",
    "create_mapped_field",
    "export_raw_data",
]

config = load_config()
logger = logging.getLogger(__name__)


# Потолок страницы для диагностических (/api/logs/*) эндпоинтов: они админские
# и читаются человеком при разборе инцидента, где выборка в 200 записей часто
# обрывается на середине проблемы. Тело записи RequestLog ограничено сверху
# (RequestLoggingMiddleware.MAX_BODY_LENGTH), так что ответ остаётся конечным.
LOG_PAGE_SIZE_MAX = 500


def _parse_page_size(request, default: int = 50, max_value: int = 200) -> int:
    """`limit` из query string -> безопасный размер страницы для Paginator.

    Невалидированный limit ронял вьюху пятисоткой: limit=0 -> ZeroDivisionError,
    limit<0 -> EmptyPage с вводящим в заблуждение текстом ("That page number is
    less than 1" — жалуется на page, хотя проблема в limit), limit=abc ->
    ValueError из int(). Клиент получал 500 с сырым текстом Python-исключения,
    а log_errors на каждый такой запрос писал в SystemLog ERROR с traceback —
    шум в мониторинге на банальной ошибке ввода.

    Верхняя граница не даёт ?limit=100000 сериализовать всю таблицу в один
    ответ; конкретное значение задаёт вызывающий эндпоинт.
    """
    try:
        return max(1, min(int(request.GET.get("limit", default)), max_value))
    except (TypeError, ValueError):
        return default


def _parse_refresh_flag(request) -> bool:
    """`?refresh=...` из query string -> bool "принудительно обойти кэш".

    Общий разбор для всех эндпоинтов с форс-рефрешем (get_project_board_meta,
    get_project_board, get_homepage_portfolio) — сравнение с множеством
    "истинных" строк, а не int()/bool(): не бросает исключений ни при каком
    значении параметра (пустая строка, слово, что угодно). План уже дважды
    ловил падения на "голом" int()/list() парсинге вне try/except (Task 1,
    fix rounds 1-2) — оба раза именно из-за того, что один и тот же параметр
    парсился по-разному в разных местах. Один разбор на все три места, а не
    копия в каждом, — чтобы не завести третий такой дефект.
    """
    return str(request.GET.get("refresh", "")).strip().lower() in {"1", "true", "y", "yes"}


def _get_filtered_timesheet_queryset(request: AuthorizedRequest):
    return build_filtered_timesheet_queryset(
        request.bitrix24_account,
        {
            "date_from": request.GET.get("date_from"),
            "date_to": request.GET.get("date_to"),
            "employee_ids[]": request.GET.getlist("employee_ids[]"),
            "project_ids[]": request.GET.getlist("project_ids[]"),
            "employee_mode": request.GET.get("employee_mode", "include"),
            "project_mode": request.GET.get("project_mode", "include"),
        },
    )


def _get_user_map(request: AuthorizedRequest, user_ids):
    """Строит {employee_id: "Фамилия Имя"}: быстрый путь — локальная БД
    (portal_user), Bitrix user.get — страховка для id, которых там нет.
    Быстрый путь убирает 3-7с "user_map" на отчётах (был холодный промах
    per-воркер Django LocMemCache) — см. Фаза 2 sync-offload.

    Входящие user_ids (из TimesheetItem.employee_id, в т.ч. историчные
    строки) нормализуются через extract_bitrix_user_id ДО запроса — тем же
    конвертером, которым UserSyncService пишет PortalUser.bitrix_id, — иначе
    неканоничные формы ("[12]", "12.0") не совпадут с каноничным bitrix_id в
    БД. Ключи результата тоже каноничные: resolve_employee_name ищет сначала
    по normalize_employee_id(employee_id), так что канонический ключ находит
    имя и для сырого, и для неканоничного значения строки.

    Хотфикс 2026-07-28 (прод-регресс «User <id>» в дереве задачи): синк
    PortalUser молодой и может не покрывать всех сотрудников (плюс отдельно
    чинится баг парсинга ACTIVE в UserSyncService, из-за которого локальный
    справочник был почти пуст) — id, которых нет в локальной БД, дорезолвятся
    через BitrixDataService.fetch_users (со своим LocMemCache). Сбой Bitrix
    здесь НЕ должен ронять отчёт: в худшем случае недостающие имена просто не
    резолвятся (fallback на "Сотрудник <id>" — уровнем выше, в
    resolve_employee_name). Данные из PortalUser всегда в приоритете —
    Bitrix запрашивается только за тем, чего не хватило локально.
    """
    if not user_ids:
        return {}

    normalized_ids = {extract_bitrix_user_id(uid) for uid in user_ids}
    normalized_ids.discard("")
    if not normalized_ids:
        return {}

    rows = PortalUser.objects.filter(
        **scope_to_tenant(request.bitrix24_account),
        bitrix_id__in=list(normalized_ids),
    ).values("bitrix_id", "name", "last_name")

    user_map = {
        row["bitrix_id"]: (f"{row['last_name']} {row['name']}".strip() or row["bitrix_id"])
        for row in rows
    }

    missing_ids = normalized_ids - user_map.keys()
    if missing_ids:
        fallback_map = {}
        try:
            fallback_map = _get_data_service(request).fetch_users(list(missing_ids))
        except Exception:
            logger.exception(
                "_get_user_map: Bitrix-фоллбэк упал для %d недостающих id", len(missing_ids)
            )
        for uid, name in fallback_map.items():
            user_map.setdefault(uid, name)
        logger.info(
            "_get_user_map: PortalUser не покрыл %d из %d id, Bitrix-фоллбэк дорезолвил %d",
            len(missing_ids), len(normalized_ids), len(fallback_map),
        )

    return user_map


def _current_user_display_name(request: AuthorizedRequest) -> str:
    """Имя текущего сотрудника для поля «куратор». Берём из локального
    справочника (PortalUser) — он же питает user_map отчётов; если сотрудника
    там ещё нет, куратор останется с пустым именем, но с корректным id."""
    account = request.bitrix24_account
    row = PortalUser.objects.filter(
        **scope_to_tenant(account),
        bitrix_id=str(account.b24_user_id or ""),
    ).values("name", "last_name").first()
    if not row:
        return ""
    return f"{row['last_name']} {row['name']}".strip()


def _get_data_service(request: AuthorizedRequest):
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()
    return BitrixDataService(request.bitrix24_account.client, config, request.bitrix24_account)


def _build_employee_filter_options(request: AuthorizedRequest):
    data_service = _get_data_service(request)
    return data_service.fetch_active_users()


def _build_project_filter_options(request: AuthorizedRequest):
    return build_project_filter_options(request.bitrix24_account)


PROJECT_SPA_REQUIRED_MAPPING = {
    "title": "string",
    "bitrix_group_id": "project_identifier",
    "stage_id": "stage",
    "is_support": "boolean",
    "project_hours_budget": "double",
    "hourly_rate": "double",
    "curator_id": "employee",
    "company_id": "crm_binding",
    "our_legal_entity_id": "crm_binding",
    "start_date": "date",
    "finish_date": "date",
    "is_archived": "boolean",
}

PROJECT_SPA_TYPE_ALIASES = {
    "string": {"string", "text", "char"},
    "integer": {"integer", "int"},
    "double": {"double", "float", "money"},
    "boolean": {"boolean", "bool"},
    "employee": {"employee", "user", "crm_status"},
    # Ключ — логическое ожидание приложения, значения — то, что реально может
    # вернуть crm.item.fields. "crm" — штатный тип привязки к элементам CRM;
    # "crm_company" остаётся ради порталов, где поле заводили руками до того, как
    # выяснилось, что такого типа в Битриксе нет; "string"/"integer" — ради тех,
    # где под компанию подобрали обычное поле с ID.
    "crm_binding": {"crm", "crm_company", "string", "integer"},
    "date": {"date", "datetime"},
    "project_identifier": {"integer", "int", "string", "text", "char"},
    "stage": {"string", "text", "char", "crm_status", "status"},
}


def _normalize_field_type(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_project_field_type_compatible(expected_type, actual_type):
    actual = _normalize_field_type(actual_type)
    expected_aliases = PROJECT_SPA_TYPE_ALIASES.get(expected_type, {expected_type})
    return actual in expected_aliases


def _collect_project_spa_linkage_issues(sync_service: ProjectSyncService, entity_type_id: int, project_mapping: dict):
    items = sync_service.fetch_project_sp_items(entity_type_id)
    normalized = [
        sync_service.normalize_project_item(item, project_mapping, {}, {}, {})
        for item in items
    ]

    duplicate_map = defaultdict(set)
    duplicate_item_map = defaultdict(set)
    missing_group_link_count = 0
    for item in normalized:
        bitrix_group_id = str(item.get("project_id") or "").strip()
        project_item_id = str(item.get("project_item_id") or "").strip()
        if not bitrix_group_id:
            missing_group_link_count += 1
            continue
        if project_item_id:
            duplicate_map[bitrix_group_id].add(project_item_id)
            duplicate_item_map[project_item_id].add(bitrix_group_id)

    duplicate_links = [
        {
            "bitrix_group_id": group_id,
            "project_item_ids": sorted(project_item_ids),
        }
        for group_id, project_item_ids in duplicate_map.items()
        if len(project_item_ids) > 1
    ]
    duplicate_item_links = [
        {
            "project_item_id": project_item_id,
            "bitrix_group_ids": sorted(group_ids),
        }
        for project_item_id, group_ids in duplicate_item_map.items()
        if len(group_ids) > 1
    ]

    return {
        "total_items": len(normalized),
        "missing_group_link_count": missing_group_link_count,
        "duplicate_group_link_count": len(duplicate_links),
        "duplicate_group_links": duplicate_links,
        "duplicate_project_item_link_count": len(duplicate_item_links),
        "duplicate_project_item_links": duplicate_item_links,
    }


def _build_project_spa_validation_payload(
    config_service: ConfigurationService,
    account,
    config: dict,
):
    config = config_service.normalize_configuration_sync(config)
    try:
        entity_type_id = int(config.get("project_sp_entity_type_id") or 0)
    except (TypeError, ValueError):
        entity_type_id = 0
    project_mapping = config.get("project_fields_mapping") or {}
    if not isinstance(project_mapping, dict):
        project_mapping = {}

    required_keys = list(PROJECT_SPA_REQUIRED_MAPPING.keys())
    missing_mapping_keys = []
    missing_fields_in_sp = []
    type_mismatches = []
    warnings = []
    access_error = None
    write_access_error = None
    linkage_issues = {
        "total_items": 0,
        "missing_group_link_count": 0,
        "duplicate_group_link_count": 0,
        "duplicate_group_links": [],
        "duplicate_project_item_link_count": 0,
        "duplicate_project_item_links": [],
    }

    if not entity_type_id:
        return {
            "is_configured": False,
            "is_valid": False,
            "entity_type_id": 0,
            "required_mapping_keys": required_keys,
            "missing_mapping_keys": required_keys,
            "missing_fields_in_sp": [],
            "type_mismatches": [],
            "access_error": "Не выбран Смарт-процесс ПРОЕКТ в настройках.",
            "write_access_error": "Не выбран Смарт-процесс ПРОЕКТ в настройках.",
            "warnings": ["Project SPA не настроен."],
            "linkage_issues": linkage_issues,
        }

    fields_meta = config_service.get_sp_fields_sync(entity_type_id)
    fields_by_id = {str(field.get("id") or ""): field for field in fields_meta if field.get("id")}
    fields_by_id_lower = {key.lower(): value for key, value in fields_by_id.items()}

    for mapping_key, expected_type in PROJECT_SPA_REQUIRED_MAPPING.items():
        mapped_field = str(project_mapping.get(mapping_key) or "").strip()
        if not mapped_field:
            missing_mapping_keys.append(mapping_key)
            continue

        field_meta = fields_by_id.get(mapped_field) or fields_by_id_lower.get(mapped_field.lower())
        if not field_meta:
            missing_fields_in_sp.append({"key": mapping_key, "mapped_field": mapped_field})
            continue

        actual_type = field_meta.get("type")
        if not _is_project_field_type_compatible(expected_type, actual_type):
            type_mismatches.append(
                {
                    "key": mapping_key,
                    "mapped_field": mapped_field,
                    "expected_type": expected_type,
                    "actual_type": _normalize_field_type(actual_type),
                }
            )

    try:
        account.client._bitrix_token.call_method(
            "crm.item.list",
            {"entityTypeId": entity_type_id, "select": ["id"], "start": 0},
        )
    except Exception as exc:
        access_error = str(exc)

    try:
        sample_items_response = account.client._bitrix_token.call_method(
            "crm.item.list",
            {"entityTypeId": entity_type_id, "select": ["id", "title"], "start": 0},
        )
        sample_items = sample_items_response.get("result", [])
        if isinstance(sample_items, dict):
            sample_items = sample_items.get("items") or sample_items.get("result") or []
        sample_id = None
        for row in sample_items or []:
            sample_id = row.get("id") or row.get("ID")
            if sample_id:
                break
        if sample_id:
            account.client._bitrix_token.call_method(
                "crm.item.update",
                {
                    "entityTypeId": entity_type_id,
                    "id": int(sample_id) if str(sample_id).isdigit() else sample_id,
                    "fields": {},
                },
            )
        else:
            warnings.append("Права на обновление не проверены: в Project SPA нет элементов для no-op update.")
    except Exception as exc:
        exc_text = str(exc)
        benign_markers = ("fields", "empty", "пуст")
        if any(marker in exc_text.lower() for marker in benign_markers):
            warnings.append("Права на обновление подтверждены частично: no-op update не принял пустой payload.")
        else:
            write_access_error = exc_text

    try:
        sync_service = ProjectSyncService(account.client, account)
        linkage_issues = _collect_project_spa_linkage_issues(sync_service, entity_type_id, project_mapping)
    except Exception as exc:
        warnings.append(f"Не удалось проверить связность Project SPA: {exc}")

    if linkage_issues["missing_group_link_count"] > 0:
        warnings.append(
            f"Есть проекты без bitrix_group_id: {linkage_issues['missing_group_link_count']}."
        )
    if linkage_issues["duplicate_group_link_count"] > 0:
        warnings.append(
            f"Есть конфликты group_id -> project_item_id: {linkage_issues['duplicate_group_link_count']}."
        )
    if linkage_issues["duplicate_project_item_link_count"] > 0:
        warnings.append(
            f"Есть конфликты project_item_id -> group_id: {linkage_issues['duplicate_project_item_link_count']}."
        )

    is_valid = (
        not missing_mapping_keys
        and not missing_fields_in_sp
        and not type_mismatches
        and access_error is None
        and write_access_error is None
    )

    return {
        "is_configured": True,
        "is_valid": is_valid,
        "entity_type_id": entity_type_id,
        "required_mapping_keys": required_keys,
        "missing_mapping_keys": missing_mapping_keys,
        "missing_fields_in_sp": missing_fields_in_sp,
        "type_mismatches": type_mismatches,
        "access_error": access_error,
        "write_access_error": write_access_error,
        "warnings": warnings,
        "linkage_issues": linkage_issues,
    }


def _load_request_json(request: AuthorizedRequest):
    """Разбирает JSON-тело запроса и ГАРАНТИРУЕТ словарь на выходе.

    Все вызывающие (create_project_board, update_project_board,
    update_project_board_stage, archive_project_board) обращаются к
    результату как к dict — сразу .get(...), без проверки типа. Раньше сюда
    пропускалось любое успешно разобранное JSON-значение как есть: null и []
    случайно гасились чужой конструкцией `x or {}` ниже по стеку (например,
    resolve_project_fields), а вот непустой список/число/строка/true —
    истинны, "x or {}" их не трогает, и они долетали до первого чужого
    payload.get(...) с 500 "'list'/'int'/'str'/'bool' object has no
    attribute 'get'". Проверка типа здесь, а не у каждого вызывающего:
    у хелпера один контракт на всех, а не N мест, которые обязаны помнить
    о нём сами.
    """
    try:
        parsed = json.loads(request.body or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _apply_created_at_filters(queryset, created_from, created_to):
    from datetime import date, datetime, time, timedelta

    current_tz = timezone.get_current_timezone()

    def parse_date(raw_value):
        if not raw_value:
            return None
        try:
            return date.fromisoformat(str(raw_value)[:10])
        except (TypeError, ValueError):
            return None

    parsed_from = parse_date(created_from)
    parsed_to = parse_date(created_to)

    if parsed_from:
        queryset = queryset.filter(created_at__gte=timezone.make_aware(datetime.combine(parsed_from, time.min), current_tz))
    if parsed_to:
        next_day = parsed_to + timedelta(days=1)
        queryset = queryset.filter(created_at__lt=timezone.make_aware(datetime.combine(next_day, time.min), current_tz))
    return queryset


def _build_export_date_filter(date_type, date_from, date_to, fields_mapping):
    """Фильтр периода для crm.item.list в выгрузке export_raw_data.

    Верхняя граница — «строго меньше начала следующего дня», как в
    _apply_created_at_filters выше. Через "<=date_to" её задавать нельзя:
    Битрикс читает строку-дату без времени как НАЧАЛО суток, и всё созданное
    в последний день периода молча выпадало из выгрузки (пользователь просит
    «по 30.07 включительно» — получает по 29.07). Для CREATED_TIME это било
    всегда, для даты отражения — если портал завёл поле как datetime.
    """
    from datetime import date as date_cls, timedelta

    # FIX: frontend sends "creation", not "created"
    if date_type == "creation":
        field = "CREATED_TIME"
    else:
        # FIX: the reflection date field key in config is "data", not "date_reflection"
        field = (fields_mapping or {}).get("data", "CREATED_TIME")

    crm_filter = {}
    if date_from:
        crm_filter[f">={field}"] = date_from
    if date_to:
        try:
            next_day = date_cls.fromisoformat(str(date_to)[:10]) + timedelta(days=1)
        except (TypeError, ValueError):
            # Неразбираемое значение отдаём Битриксу как есть — это его 400,
            # а не наш 500 на разборе пользовательского ввода.
            crm_filter[f"<={field}"] = date_to
        else:
            crm_filter[f"<{field}"] = next_day.isoformat()
    return crm_filter


def _serialize_support_status(payload):
    return {
        "configured": bool(payload.get("configured")),
        "code": payload.get("code") or "",
        "status": payload.get("status") or "not_connected",
        "dialog_id": payload.get("dialog_id") or "",
        "connected_at": payload.get("connected_at"),
        "error": payload.get("error") or "",
    }


@xframe_options_exempt
@require_GET
@log_errors("root")
@auth_required
def root(request: AuthorizedRequest):
    return JsonResponse({"message": "Python Backend is running"})


@xframe_options_exempt
@require_GET
@log_errors("health")
@auth_required
def health(request: AuthorizedRequest):
    return JsonResponse({
        "status": "healthy",
        "backend": "python",
        "timestamp": timezone.now().timestamp(),
    })

def health_check(request):
    """
    Public health check for deployment platforms (e.g. Timeweb Cloud Apps).
    No auth required.
    """
    return JsonResponse({
        "status": "ok",
        "timestamp": timezone.now().timestamp(),
    })



@xframe_options_exempt
@require_GET
@log_errors("get_enum")
@auth_required
def get_enum(request: AuthorizedRequest):
    options = ["option 1", "option 2", "option 3"]
    return JsonResponse(options, safe=False)


@xframe_options_exempt
@require_GET
@log_errors("get_list")
@auth_required
def get_list(request: AuthorizedRequest):
    elements = ["element 1", "element 2", "element 3"]
    return JsonResponse(elements, safe=False)


# Как часто сверять флаг администратора с Bitrix. Права администратора портала
# меняются крайне редко, а сверка — блокирующий REST-вызов в критическом пути
# выдачи токена, см. _refresh_admin_flag.
ADMIN_FLAG_TTL = timedelta(hours=6)


def _refresh_admin_flag(account, force: bool = False) -> None:
    """Refresh ``account.is_b24_user_admin`` from the Bitrix ``user.admin`` method.

    Вызывается из install (``force=True``) и из выдачи токена. Раньше сверка шла
    на КАЖДЫЙ /api/getToken, а initApp() стоит в onMounted каждой страницы — то
    есть на открытие вкладки задачи приходилось два блокирующих вызова Bitrix
    подряд, каждый из которых ещё и занимал слот gunicorn (их всего 8) и жёг
    общий на портал операционный лимит метода user.admin.

    Теперь сверка идёт не чаще раза в ADMIN_FLAG_TTL. При ``force=True`` (установка
    приложения) сверяем всегда — там флаг заводится впервые.

    Сбой вызова Bitrix НЕ должен ломать аутентификацию: прежнее значение
    сохраняется, пишется только предупреждение. Отметка времени при сбое не
    обновляется — иначе неудачная сверка «залипала» бы на весь TTL.
    """
    if not force and account.admin_flag_checked_at is not None:
        if timezone.now() - account.admin_flag_checked_at < ADMIN_FLAG_TTL:
            return

    try:
        response = account.client._bitrix_token.call_method("user.admin", {})
        is_admin = bool(response.get("result") if isinstance(response, dict) else response)

        updated_fields = ["admin_flag_checked_at"]
        if account.is_b24_user_admin != is_admin:
            account.is_b24_user_admin = is_admin
            updated_fields.append("is_b24_user_admin")

        account.admin_flag_checked_at = timezone.now()
        account.save(update_fields=updated_fields)
    except Exception:
        logger.warning("Could not refresh is_b24_user_admin via user.admin for account %s", account.pk, exc_info=True)


@xframe_options_exempt
@csrf_exempt
@log_errors("install")
def install(request):
    """
    Handle Bitrix24 application installation.
    Supports HEAD/GET for Marketplace validation.
    """
    logger.info("Install view called. Method=%s Path=%s", request.method, request.path)
    if request.method in ["HEAD", "GET"]:
        return HttpResponse(status=200)

    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    # Manual Auth Check since we removed @auth_required to support HEAD
    try:
        # Re-using the decorator logic or manually calling the auth service
        # Since @auth_required populated request.bitrix24_account, we need to do it here manually
        # effectively inlining the auth check for POST requests ONLY.
        
        # However, to keep it clean, let's keep @auth_required but make it smarter?
        # No, decorators are hard to make conditional on method easily without complexity.
        # Simplest way: wrapper function.
        pass
    except Exception:
        pass

    # Better approach: 
    # We can't easily inline the complex auth logic from decorators.
    # Let's use a dual-handler approach or simply wrap the logic.
    
    return _install_post_logic(request)

@require_POST
@auth_required
def _install_post_logic(request: AuthorizedRequest):
    bitrix24_account = request.bitrix24_account

    # На установке сверяем всегда: аккаунт заводится впервые, TTL ещё не с чем сравнивать.
    _refresh_admin_flag(bitrix24_account, force=True)

    ApplicationInstallation.objects.update_or_create(
        bitrix_24_account=bitrix24_account,
        defaults={
            "status": bitrix24_account.status,
            "portal_license_family": "",
            "application_token": bitrix24_account.application_token,
        },
    )

    try:
        service = InstallationService(bitrix24_account.client, bitrix24_account)
        config = service.install_app_sync()
        support = _serialize_support_status(service.get_support_line_status())
        return JsonResponse({"message": "Installation successful", "config": config, "support": support})
    except InstallationError as e:
        return JsonResponse({"error": str(e)}, status=500)
    except Exception:
        logger.exception("Unexpected error during installation")
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("get_token")
@rate_limit("get_token", 10, 60, key="ip_domain")
@auth_required
def get_token(request: AuthorizedRequest):
    _refresh_admin_flag(request.bitrix24_account)
    return JsonResponse({"token": request.bitrix24_account.create_jwt_token()})


@xframe_options_exempt
@require_GET
@log_errors("get_filter_options")
@auth_required
def get_filter_options(request: AuthorizedRequest):
    """
    Returns unique employees and projects for filtering.
    """
    return JsonResponse({
        "employees": _build_employee_filter_options(request),
        "projects": _build_project_filter_options(request)
    })


@xframe_options_exempt
@require_GET
@log_errors("get_filter_employees")
@auth_required
def get_filter_employees(request: AuthorizedRequest):
    return JsonResponse({"employees": _build_employee_filter_options(request)})


@xframe_options_exempt
@require_GET
@log_errors("get_filter_projects")
@auth_required
def get_filter_projects(request: AuthorizedRequest):
    return JsonResponse({"projects": _build_project_filter_options(request)})


@xframe_options_exempt
@require_GET
@log_errors("get_support_status")
@auth_required
def get_support_status(request: AuthorizedRequest):
    service = InstallationService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(_serialize_support_status(service.get_support_line_status()))


@xframe_options_exempt
@require_POST
@log_errors("connect_support_line")
@auth_required
def connect_support_line(request: AuthorizedRequest):
    service = InstallationService(request.bitrix24_account.client, request.bitrix24_account)
    payload = _serialize_support_status(service.connect_support_line_sync(force=True))
    return JsonResponse(payload)


@xframe_options_exempt
@require_GET
@log_errors("get_project_board")
@auth_required
def get_project_board(request: AuthorizedRequest):
    """?refresh=1 — принудительный обход серверного кэша доски (2 минуты,

    PROJECT_BOARD_CACHE_TTL). Тот же приём, что и у get_project_board_meta
    (см. её докстринг про разбор параметра): нужен, потому что кэш —
    LocMemCache, свой у каждого воркера gunicorn. invalidate_project_runtime_caches,
    которую create() зовёт после write_through, чистит кэш только того
    воркера, который обработал запрос на создание; следующий GET доски может
    уйти на другой воркер и получить кэш, прогретый до создания проекта, —
    см. докстринг ProjectCardService.get_board_data.

    БЕЗ @rate_limit, в отличие от board_meta_refresh: там форс-рефреш бьёт
    живьём в Битрикс (app.option.get + crm.company.list), а тут — нет.
    get_board_data(bypass_cache=True) пропускает ТОЛЬКО чтение кэша
    "project-board" и пересчитывает ответ из локальной базы; единственный
    живой вызов Битрикса на этом пути (app.option.get в _load_config)
    происходит и без всякого refresh — на любой органический холодный кэш,
    то есть и так не реже раза в 2 минуты на аккаунт без всякого лимита
    сегодня. refresh=1 просто просит совершить этот же пересчёт по запросу
    клиента, а не по истечении TTL — лимитировать только эту ветку, оставляя
    безлимитным обычный путь с тем же наихудшим темпом, не защитило бы
    бюджет Битрикса и сломало бы легитимный сценарий (перечитать доску сразу
    после нажатия «Создать проект»).
    """
    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(service.get_board_data(bypass_cache=_parse_refresh_flag(request)))


@rate_limit("board_meta_refresh", 6, 60, key="account")
def _get_project_board_meta_refresh(request: AuthorizedRequest, service: ProjectCardService) -> JsonResponse:
    """Лимитированная ветка ?refresh=1 — вынесена отдельно, чтобы @rate_limit

    покрывал только её. Только при bypass_cache=True get_meta реально ходит в
    Битрикс живьём: app.option.get в _load_config (кэш ConfigurationService —
    на объекте, новый объект создаётся на каждый вызов, поэтому не переживает
    его) и crm.company.list в get_legal_entities/list_my_companies. Обычные
    запросы (ветка ниже, в get_project_board_meta) отдаются из серверного
    кэша project-board-meta (TTL 6 часов) и Битрикс не трогают вовсе — им
    лимит не нужен и он их не касается.

    Порог 6/60 — тот же класс риска и то же число, что у соседнего
    sync_project_board (@rate_limit("sync", 6, 60, key="account")): там же
    ровно 1-2 живых вызова к Bitrix за запрос. Отдельный scope
    ("board_meta_refresh", а не "sync") — иначе один клик «Синхронизировать»
    (фронт бьёт и /project-board/sync, и /project-board/meta?refresh=1 одним
    действием, см. frontend/app/pages/projects/index.client.vue:syncBoard)
    тратил бы бюджет обоих эндпоинтов из одного и того же ведра и не давал
    бы затем ещё и нажать «Обновить справочники» отдельно.
    """
    return JsonResponse(service.get_meta(bypass_cache=True))


@xframe_options_exempt
@require_GET
@log_errors("get_project_board_meta")
@auth_required
def get_project_board_meta(request: AuthorizedRequest):
    # ?refresh=1 — кнопка «Обновить справочники» (принудительный обход
    # серверного кэша project-board-meta, живущего 6 часов). Разбор параметра
    # вынесен в _parse_refresh_flag — общий для этого эндпоинта и его соседей
    # get_project_board/get_homepage_portfolio (см. её докстринг).
    bypass_cache = _parse_refresh_flag(request)
    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)
    if bypass_cache:
        # Только эта ветка бьёт в Битрикс живьём — см. docstring
        # _get_project_board_meta_refresh. Обычные запросы (ниже) читаются
        # из кэша и вызываются с доски часто и штатно; лимитировать их вместе
        # с refresh сломало бы обычную работу доски.
        return _get_project_board_meta_refresh(request, service)
    return JsonResponse(service.get_meta(bypass_cache=False))


@xframe_options_exempt
@require_GET
@log_errors("get_project_board_card")
@auth_required
def get_project_board_card(request: AuthorizedRequest):
    project_id = str(request.GET.get("project_id") or "").strip()
    if not project_id:
        return JsonResponse({"error": "project_id is required"}, status=400)

    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)

    try:
        return JsonResponse({"card": service.get_card_data(project_id)})
    except ProjectCard.DoesNotExist:
        # "Карточки ещё нет" — не ошибка сервера, а валидное состояние:
        # для задач/групп без синхронизированного ProjectCard placement просто не показывает блок.
        # Возвращаем 200 с null, чтобы ofetch на фронте не бросал FetchError и не блокировал
        # нативное списание времени в Битрикс24.
        return JsonResponse({"card": None})
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=400)


@xframe_options_exempt
@require_GET
@log_errors("get_project_board_companies")
@auth_required
def get_project_board_companies(request: AuthorizedRequest):
    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse({"companies": service.get_companies()})


@xframe_options_exempt
@require_GET
@log_errors("search_project_board_companies")
@auth_required
@rate_limit("company_search", 60, 60, key="account")
def search_project_board_companies(request: AuthorizedRequest):
    """Поиск компаний по мере ввода. Полный справочник портала (десятки тысяч
    записей на боевом) сюда не выгружается — см. company_search_service.py.

    `limit` передаётся в CompanySearchService.search как есть, без разбора на
    уровне view: `_parse_limit` внутри сервиса уже переживает любой мусор —
    пустую строку («?limit=» без значения — request.GET.get отдаёт "", а не
    None), нечисловые значения, отсутствие параметра вовсе. Дублировать эту
    логику здесь не стоит: в этом плане уже дважды ловили дефекты именно
    из-за разъехавшегося разбора одного и того же параметра в двух местах.

    В отличие от ?refresh=1 у get_project_board_meta, здесь дешёвой ветки нет:
    CompanySearchService.search всегда бьёт crm.company.list живьём, а для
    похожих на ИНН запросов — ещё и crm.requisite.list (см. company_search_service.py).
    Кэш сервиса (5 минут) ключуется по точной паре «запрос+limit», так что при
    посимвольном вводе почти каждое нажатие клавиши даёт новый ключ кэша и
    новый живой вызов — задержка на фронте перед отправкой запроса лишь
    вежливость клиента, не гарантия сервера: с валидным токеном её обходит
    любой скрипт в цикле. Поэтому лимитирован весь эндпоинт, а не его часть.

    Порог — 60 запросов/60 секунд, а НЕ 6/60, как у sync/export/
    board_meta_refresh. Это не опечатка и не небрежность: у соседей другой
    профиль использования — одна ручная кнопка, нажимаемая нечасто, 6/60 для
    неё щедрый запас. У поиска профиль принципиально другой — это автокомплит:
    человек, печатающий «Ромашка» с фронтовой задержкой ~300 мс между
    запросами, легитимно порождает 2-3 живых вызова на одно слово, а заполняя
    форму, ищет подряд несколько контрагентов — десятки запросов в минуту
    совершенно законной работы. Порог 6/60 сломал бы её уже секунд через
    двадцать. 60/60 — это около одного запроса в секунду при непрерывном
    опросе: заведомо выше темпа ручного набора (даже без учёта фронтового
    debounce) и заведомо ниже того, что даст скрипт, отправляющий запросы
    подряд без пауз.

    Счётчик — свой, scope "company_search", не общий с соседями: у sync/
    export/board_meta_refresh другой профиль и другой бюджет (6/60), делить
    его было бы неверно в обе стороны — всплеск легитимного поиска не должен
    съедать чужой бюджет, и наоборот.
    """
    service = CompanySearchService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(
        service.search(request.GET.get("q") or "", limit=request.GET.get("limit"))
    )


@xframe_options_exempt
@require_GET
@log_errors("list_my_companies")
@auth_required
def list_my_companies(request: AuthorizedRequest):
    """Свои юрлица (серверный фильтр IS_MY_COMPANY вместо обхода всего
    справочника портала) — для форм, которым нужен список без полного
    справочника компаний. См. company_search_service.py."""
    service = CompanySearchService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(service.list_my_companies())


@xframe_options_exempt
@require_GET
@log_errors("get_homepage_portfolio")
@auth_required
def get_homepage_portfolio(request: AuthorizedRequest):
    """?refresh=1 — принудительный обход серверного кэша главного экрана (2

    минуты, HOMEPAGE_CACHE_TTL). Тот же приём и то же обоснование отсутствия
    @rate_limit, что и у get_project_board — см. её докстринг и докстринг
    ProjectCardService.get_homepage_snapshot (bypass_cache пробрасывается и
    во вложенный кэш доски, не только в свой собственный).
    """
    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(service.get_homepage_snapshot(bypass_cache=_parse_refresh_flag(request)))


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("sync_project_board")
@auth_required
@rate_limit("sync", 6, 60, key="account")
@sync_lock("project")
def sync_project_board(request: AuthorizedRequest):
    service = ProjectSyncService(request.bitrix24_account.client, request.bitrix24_account)
    incremental_raw = request.GET.get("incremental_since_minutes")
    incremental_since_minutes = None
    if incremental_raw:
        try:
            incremental_since_minutes = int(incremental_raw)
        except (TypeError, ValueError):
            return JsonResponse({"error": "incremental_since_minutes must be integer"}, status=400)
    try:
        return JsonResponse(service.sync(incremental_since_minutes=incremental_since_minutes))
    except Exception:
        logger.exception("Project board sync failed for account %s", request.bitrix24_account.pk)
        return JsonResponse(
            {
                "status": "warning",
                "sync_mode": "failed",
                "synced": 0,
                "created": 0,
                "updated": 0,
                "skipped_missing_group_link": 0,
                "skipped_conflict_linking": 0,
                "warning": (
                    "Синхронизацию проектов выполнить не удалось. "
                    "Показаны последние сохраненные данные."
                ),
            }
        )


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("run_project_spa_backfill")
@auth_required
def run_project_spa_backfill(request: AuthorizedRequest):
    service = ProjectSyncService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(service.backfill_timesheet_project_items())


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("create_project_board")
@auth_required
@rate_limit("project_create", 5, 60, key="account")
def create_project_board(request: AuthorizedRequest):
    """Создаёт связку «компания (+ реквизит с ИНН для новой) + группа в
    Задачах + карточка смарт-процесса».

    Шаги идемпотентны, поэтому повтор того же запроса досоздаёт только
    недостающее — фронт этим и пользуется в кнопке «Повторить».

    Исключения из service.create() наружу не протекают — оркестратор сам
    ловит все свои ошибки и возвращает частичный результат (status="error"
    по нужному шагу, done=False), поэтому здесь нет try/except: любая обёртка
    превратила бы такой частичный результат в пятисотую, а он должен доехать
    до фронта как есть, один в один.

    Порог @rate_limit — 5 запросов/60 секунд, а НЕ 6/60, как у соседних
    ручных кнопок sync/board_meta_refresh, и тем более не 60/60, как у
    company_search. Число выбрано осознанно ниже, а не переиспользовано:

    - Стоимость запроса выше, чем у sync (1-2 живых вызова, см. докстринг
      _get_project_board_meta_refresh). Здесь четыре последовательных шага
      (компания, реквизит, группа, карточка) в service.create(), и каждый
      поисковый — сначала поиск, затем, если не нашёл, создание: до двух
      живых вызовов на шаг для компании/группы/карточки. Шаг реквизита
      (только при создании НОВОЙ компании — inn-brief.md) дороже: поиск по
      ИНН, идемпотентная проверка, чтение шаблона (обычно из кэша —
      REQUISITE_PRESET_CACHE_TTL в project_creation_service.py, 6 часов) и
      создание — до четырёх живых вызовов. Итого до полутора десятков вызовов
      только на сами шаги в холодном (без кэша шаблона) худшем случае, плюс
      ещё до трёх чтений на общие для аккаунта справочники (конфигурация,
      свои юрлица, стадии проекта) — те обычно уже тёплые в кэше от обычной
      работы доски, но гарантии на это нет (напрямую в create() не
      передаются, читаются внутри с нуля).
    - Риск выше, чем у sync и чем у company_search: это не чтение, а запись
      в CRM клиента. У company_search (60/60) — это автокомплит, и его
      результат ни на что не влияет, кроме самой формы; здесь неудачный
      повтор — это реальная сущность на портале клиента (идемпотентность
      ensure_company/ensure_requisite/ensure_group спасает от дублей по
      имени/ИНН при штатном повторе, но не отменяет того, что каждое
      нажатие — это живая мутация чужого CRM, а не безобидный лишний GET).
    - Профиль использования — противоположность автокомплиту: по продукту
      это осознанное редкое действие, счёт идёт на единицы в день, а не на
      десятки в минуту (в отличие от печати запроса в поиске компаний).
      5/60 с большим запасом покрывает легитимный всплеск — первую попытку
      плюс несколько ручных нажатий кнопки «Повторить» после временного
      сбоя (лок занят, сетевая заминка) — и при этом останавливает скрипт,
      бьющий по эндпоинту в цикле, уже в первую минуту.

    Отдельный scope "project_create" (см. tests_security_ratelimit.py) — не
    общий со sync/company_search/остальными. Имя совпадает со
    scope="project_create" у account_sync_lock в ProjectCreationService.create()
    (тот же логический ярлык для одного и того же действия), но механизмы
    независимы: Django-кэш с префиксом "rl:" здесь, Postgres advisory-lock
    там, — общего пространства ключей нет.
    """
    payload = _load_request_json(request)
    account = request.bitrix24_account

    service = ProjectCreationService(account.client, account)
    result = service.create(
        payload,
        current_user_id=str(account.b24_user_id or ""),
        current_user_name=_current_user_display_name(request),
    )
    return JsonResponse(result)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("update_project_board")
@auth_required
def update_project_board(request: AuthorizedRequest):
    payload = _load_request_json(request)
    project_id = payload.get("project_id")
    if not project_id:
        return JsonResponse({"error": "project_id is required"}, status=400)

    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)

    try:
        result = service.update_project_card(str(project_id), payload)
        return JsonResponse(result)
    except ProjectCard.DoesNotExist:
        return JsonResponse({"error": "Проект не найден"}, status=404)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=400)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("update_project_board_stage")
@auth_required
def update_project_board_stage(request: AuthorizedRequest):
    payload = _load_request_json(request)
    project_id = payload.get("project_id")
    stage = payload.get("stage")

    if not project_id or not stage:
        return JsonResponse({"error": "project_id and stage are required"}, status=400)

    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)

    try:
        result = service.update_stage(str(project_id), str(stage))
        return JsonResponse(result)
    except ProjectCard.DoesNotExist:
        return JsonResponse({"error": "Проект не найден"}, status=404)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=400)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("archive_project_board")
@auth_required
def archive_project_board(request: AuthorizedRequest):
    payload = _load_request_json(request)
    project_id = payload.get("project_id")
    is_archived = str(payload.get("is_archived", "")).strip().lower() in {"1", "true", "y", "yes", "on"}

    if not project_id:
        return JsonResponse({"error": "project_id is required"}, status=400)

    service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)

    try:
        result = service.archive_project(str(project_id), is_archived)
        return JsonResponse(result)
    except ProjectCard.DoesNotExist:
        return JsonResponse({"error": "Проект не найден"}, status=404)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=400)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("run_project_board_daily_check")
@auth_required
def run_project_board_daily_check(request: AuthorizedRequest):
    service = ProjectStageAutomationService(request.bitrix24_account)
    return JsonResponse(service.run_daily_check())


@xframe_options_exempt
@require_GET
@log_errors("report_employee_project")
@auth_required
def report_employee_project(request: AuthorizedRequest):
    profiler = ReportProfiler("report_employee_project", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("materialize"):
        rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)
    with profiler.stage("project_lookup"):
        project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    with profiler.stage("build_items"):
        items = build_tree_report_items(
            rows,
            project_name_by_item=project_name_by_item,
            project_name_by_group=project_name_by_group,
        )
    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_employee_projects(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report, safe=False)
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_project_employee")
@auth_required
def report_project_employee(request: AuthorizedRequest):
    profiler = ReportProfiler("report_project_employee", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("materialize"):
        rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)
    with profiler.stage("project_lookup"):
        project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    with profiler.stage("build_items"):
        items = build_tree_report_items(
            rows,
            project_name_by_item=project_name_by_item,
            project_name_by_group=project_name_by_group,
        )
    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_project_employees(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report, safe=False)
    profiler.attach_to_response(response)
    profiler.log()
    return response

@xframe_options_exempt
@require_GET
@log_errors("report_project_task_employee")
@auth_required
def report_project_task_employee(request: AuthorizedRequest):
    """Report: Project -> Task Hierarchy -> Employee -> Items"""
    profiler = ReportProfiler("report_project_task_employee", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("materialize"):
        rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)
    with profiler.stage("project_lookup"):
        project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    with profiler.stage("build_items"):
        items = build_tree_report_items(
            rows,
            include_task_id=True,
            project_name_by_item=project_name_by_item,
            project_name_by_group=project_name_by_group,
        )
    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_project_task_employees(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report, safe=False)
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_project_task_employee_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_project_task_employee_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «Учет по проектам/задачам» с сохранением иерархии."""
    queryset = _get_filtered_timesheet_queryset(request)
    rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    user_map = _get_user_map(request, user_ids)
    project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    items = build_tree_report_items(
        rows,
        include_task_id=True,
        project_name_by_item=project_name_by_item,
        project_name_by_group=project_name_by_group,
    )
    report = ReportService().generate_project_task_employees(items, user_map)

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_project_task_workbook(report, date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    suffix = f"_{date_from}_{date_to}".strip("_")
    filename = f"report_project_task{('_' + suffix) if suffix else ''}.xlsx".replace("__", "_")
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_employee_project_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_employee_project_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «По сотрудникам/проектам» с сохранением иерархии."""
    queryset = _get_filtered_timesheet_queryset(request)
    rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    user_map = _get_user_map(request, user_ids)
    project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    items = build_tree_report_items(
        rows,
        project_name_by_item=project_name_by_item,
        project_name_by_group=project_name_by_group,
    )
    roots = ReportService().generate_employee_projects(items, user_map)

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_hierarchy_workbook(roots, title="Отчет по сотрудникам",
                                          date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_employee_project.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_project_employee_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_project_employee_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «По проектам/сотрудникам» с сохранением иерархии."""
    queryset = _get_filtered_timesheet_queryset(request)
    rows = materialize_rows(queryset, TREE_REPORT_FIELDS)
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    user_map = _get_user_map(request, user_ids)
    project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    items = build_tree_report_items(
        rows,
        project_name_by_item=project_name_by_item,
        project_name_by_group=project_name_by_group,
    )
    roots = ReportService().generate_project_employees(items, user_map)

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_hierarchy_workbook(roots, title="Отчет по проектам",
                                          date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_project_employee.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_daily_workload_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_daily_workload_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «Ежедневная нагрузка» в виде матрицы сотрудник×день."""
    from datetime import date
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from:
        today = date.today()
        date_from = date(today.year, today.month, 1).isoformat()
    if not date_to:
        today = date.today()
        date_to = today.isoformat()

    queryset = build_filtered_timesheet_queryset(
        request.bitrix24_account,
        {
            "date_from": date_from,
            "date_to": date_to,
            "employee_ids[]": request.GET.getlist("employee_ids[]"),
            "project_ids[]": request.GET.getlist("project_ids[]"),
            "employee_mode": request.GET.get("employee_mode", "include"),
            "project_mode": request.GET.get("project_mode", "include"),
        },
    )
    project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    rows = materialize_rows(
        queryset,
        (
            "employee_id",
            "project_item_id",
            "project_id",
            "project_title",
            "hours",
            "task_id",
            "task_hierarchy_titles",
            "description",
            "date_reflection",
        ),
    )
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    user_map = _get_user_map(request, user_ids)
    items = [
        {
            "sotrudnik_id": row["employee_id"],
            "project_name": resolve_project_name_for_row(row, project_name_by_item, project_name_by_group),
            "kolichestvo_chasov": row["hours"],
            "id_zadachi": row["task_id"],
            "nazvanie_zadachi": row["task_hierarchy_titles"][-1] if row.get("task_hierarchy_titles") else "No Title",
            "opisanie": row["description"],
            "data": row["date_reflection"].isoformat() if row.get("date_reflection") else None,
        }
        for row in rows
    ]

    report = ReportService().generate_daily_workload(items, user_map, date_from, date_to)
    try:
        output = build_matrix_workbook(report["header_days"], report["rows"],
                                       title="Ежедневная нагрузка",
                                       date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_daily_workload.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_revenue_leakage_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_revenue_leakage_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «Потери выручки» в виде таблицы."""
    queryset = _get_filtered_timesheet_queryset(request)
    project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    rows = list(queryset.values(
        'employee_id',
        'project_item_id',
        'project_id',
        'project_title',
        'hours',
        'is_billable',
    ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    user_map = _get_user_map(request, user_ids)

    items = [{
        "sotrudnik_id": row["employee_id"],
        "project_name": resolve_project_name_for_row(row, project_name_by_item, project_name_by_group),
        "kolichestvo_chasov": row["hours"],
        "uchitivaem": row["is_billable"],
    } for row in rows]

    report = ReportService().generate_revenue_leakage(items, user_map)

    # Выгружаем risk_rows (детальные данные по рискам)
    table_rows = report.get("risk_rows", [])
    columns = [
        {"key": "project_name", "label": "Проект", "fmt": "text", "width": 25},
        {"key": "employee_name", "label": "Сотрудник", "fmt": "text", "width": 20},
        {"key": "total_hours", "label": "Всего часов", "fmt": "hours", "width": 14},
        {"key": "billable_hours", "label": "Учтено часов", "fmt": "hours", "width": 14},
        {"key": "non_billable_hours", "label": "Не учтено часов", "fmt": "hours", "width": 14},
        {"key": "loss_rate", "label": "Доля потерь, %", "fmt": "hours", "width": 14},
    ]

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_table_workbook(columns, table_rows, title="Потери выручки",
                                      date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_revenue_leakage.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_time_entry_discipline_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_time_entry_discipline_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «Дисциплина внесения времени» в виде таблицы."""
    queryset = _get_filtered_timesheet_queryset(request)
    rows = list(queryset.values(
        'employee_id',
        'date_reflection',
        'source_created_at',
        'created_at',
    ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    user_map = _get_user_map(request, user_ids)

    items = [{
        "sotrudnik_id": row["employee_id"],
        "date_reflection": row["date_reflection"],
        "source_created_at": row["source_created_at"],
        "created_at": row["created_at"],
    } for row in rows]

    report = ReportService().generate_time_entry_discipline(items, user_map)

    # Выгружаем employee_rows (детальные данные по сотрудникам)
    table_rows = report.get("employee_rows", [])
    columns = [
        {"key": "employee_name", "label": "Сотрудник", "fmt": "text", "width": 20},
        {"key": "entry_count", "label": "Записей", "fmt": "int", "width": 12},
        {"key": "same_day_share", "label": "В день записи, доля", "fmt": "percent", "width": 16},
        {"key": "avg_lag_days", "label": "Средн. отставание, дней", "fmt": "hours", "width": 16},
        {"key": "late_entries", "label": "Запис. >1дня назад", "fmt": "int", "width": 14},
        {"key": "max_lag_days", "label": "Макс. отставание, дней", "fmt": "int", "width": 16},
        {"key": "risk_level", "label": "Уровень риска", "fmt": "text", "width": 14},
    ]

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_table_workbook(columns, table_rows, title="Дисциплина внесения времени",
                                      date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_time_entry_discipline.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_focus_analysis_export")
@auth_required
@rate_limit("export", 12, 60, key="account")
def report_focus_analysis_export(request: AuthorizedRequest):
    """Excel-выгрузка отчёта «Фокус и распыление» в виде таблицы."""
    queryset = _get_filtered_timesheet_queryset(request)
    rows = list(queryset.values(
        'employee_id',
        'project_title',
        'task_id',
        'hours',
    ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    user_map = _get_user_map(request, user_ids)

    items = [{
        "sotrudnik_id": row["employee_id"],
        "project_name": row["project_title"],
        "task_id": row["task_id"],
        "kolichestvo_chasov": row["hours"],
    } for row in rows]

    report = ReportService().generate_focus_analysis(items, user_map)

    # Выгружаем employee_rows (детальные данные по сотрудникам)
    table_rows = report.get("employee_rows", [])
    columns = [
        {"key": "employee_name", "label": "Сотрудник", "fmt": "text", "width": 20},
        {"key": "project_count", "label": "Кол-во проектов", "fmt": "int", "width": 14},
        {"key": "task_count", "label": "Кол-во задач", "fmt": "int", "width": 14},
        {"key": "entry_count", "label": "Записей", "fmt": "int", "width": 12},
        {"key": "total_hours", "label": "Всего часов", "fmt": "hours", "width": 14},
        {"key": "avg_entry_hours", "label": "Средняя запись, ч", "fmt": "hours", "width": 14},
        {"key": "focus_index", "label": "Индекс фокуса", "fmt": "percent", "width": 14},
        {"key": "top_project_hours", "label": "Часов на топ проект", "fmt": "hours", "width": 16},
        {"key": "risk_level", "label": "Уровень риска", "fmt": "text", "width": 14},
    ]

    date_from = request.GET.get("date_from") or ""
    date_to = request.GET.get("date_to") or ""
    try:
        output = build_table_workbook(columns, table_rows, title="Фокус и распыление",
                                      date_from=date_from, date_to=date_to)
    except ExportTooLargeError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="report_focus_analysis.xlsx"'
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_revenue_leakage")
@auth_required
def report_revenue_leakage(request: AuthorizedRequest):
    profiler = ReportProfiler("report_revenue_leakage", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("project_lookup"):
        project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    with profiler.stage("materialize"):
        rows = list(queryset.values(
            'employee_id',
            'project_item_id',
            'project_id',
            'project_title',
            'hours',
            'is_billable',
        ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)

    with profiler.stage("build_items"):
        items = [{
            "sotrudnik_id": row["employee_id"],
            "project_name": resolve_project_name_for_row(row, project_name_by_item, project_name_by_group),
            "kolichestvo_chasov": row["hours"],
            "uchitivaem": row["is_billable"],
        } for row in rows]

    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_revenue_leakage(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report)
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_time_entry_discipline")
@auth_required
def report_time_entry_discipline(request: AuthorizedRequest):
    profiler = ReportProfiler("report_time_entry_discipline", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("materialize"):
        rows = list(queryset.values(
            'employee_id',
            'date_reflection',
            'source_created_at',
            'created_at',
        ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)

    with profiler.stage("build_items"):
        items = [{
            "sotrudnik_id": row["employee_id"],
            "date_reflection": row["date_reflection"],
            "source_created_at": row["source_created_at"],
            "created_at": row["created_at"],
        } for row in rows]

    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_time_entry_discipline(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report)
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("report_focus_analysis")
@auth_required
def report_focus_analysis(request: AuthorizedRequest):
    profiler = ReportProfiler("report_focus_analysis", account_id=request.bitrix24_account.pk)
    with profiler.stage("queryset_build"):
        queryset = _get_filtered_timesheet_queryset(request)
    with profiler.stage("materialize"):
        rows = list(queryset.values(
            'employee_id',
            'project_title',
            'task_id',
            'hours',
        ))

    user_ids = {row['employee_id'] for row in rows if row.get('employee_id')}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)

    with profiler.stage("build_items"):
        items = [{
            "sotrudnik_id": row["employee_id"],
            "project_name": row["project_title"],
            "task_id": row["task_id"],
            "kolichestvo_chasov": row["hours"],
        } for row in rows]

    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_focus_analysis(items, user_map)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report)
    profiler.attach_to_response(response)
    profiler.log()
    return response


TIMESHEET_SYNC_GATE_MINUTES = 3


def should_skip_timesheet_sync(account, now, gate_minutes=TIMESHEET_SYNC_GATE_MINUTES):
    """Гейт свежести: True, если последний синк моложе gate_minutes минут.

    account.last_timesheet_synced_at is None -> никогда не синкали -> False (синкать).
    """
    last = account.last_timesheet_synced_at
    if last is None:
        return False
    return (now - last).total_seconds() < gate_minutes * 60


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("timesheet_sync")
@auth_required
@rate_limit("sync", 6, 60, key="account")
@sync_lock("timesheet")
def timesheet_sync(request: AuthorizedRequest):
    profiler = ReportProfiler("timesheet_sync", account_id=request.bitrix24_account.pk)
    with profiler.stage("config"):
        config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
        config = config_service.get_configuration_sync()

    # Читаем необязательные даты из тела запроса (scoped/отчётный путь)
    try:
        body = json.loads(request.body or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        body = {}
    # json.loads может успешно разобрать НЕ-объект (список/число/строку/true) —
    # тогда body.get(...) ниже упал бы AttributeError. date_from/date_to здесь
    # необязательное сужение по датам, поэтому не-словарное тело трактуем как
    # "не задано" (как и при {} или битом JSON выше), а не как ошибку клиента.
    if not isinstance(body, dict):
        body = {}
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    is_scoped = bool(date_from and date_to)

    # started_at: снимается ДО обхода (спека 2026-07-31, §4.3) и только им
    # двигается маркер ниже — правка, случившаяся во время обхода, имеет
    # updatedTime >= started_at и попадёт в следующую выборку.
    now = timezone.now()
    if not is_scoped and should_skip_timesheet_sync(request.bitrix24_account, now):
        db_count = TimesheetItem.objects.filter(**scope_to_tenant(request.bitrix24_account)).count()
        profiler.set_metric("status", "fresh")
        profiler.log()
        return JsonResponse({
            "status": "fresh",
            "count": db_count,
            "last_synced_at": request.bitrix24_account.last_timesheet_synced_at.isoformat(),
        })

    # Режим синка (спека 2026-07-31, §4.1) вычисляем здесь же, чтобы решить
    # судьбу refresh_writeoff_stats: сам выбор дублируется внутри sync_all.
    mode = resolve_sync_mode(
        marker=request.bitrix24_account.last_timesheet_synced_at,
        date_from=date_from,
        date_to=date_to,
    )
    profiler.set_metric("mode", mode)
    service = TimesheetSyncService(request.bitrix24_account.client, request.bitrix24_account, config)
    try:
        with profiler.stage("sync_all"):
            count = service.sync_all(date_from=date_from, date_to=date_to)
        # refresh_writeoff_stats нужен доске проектов, не отчёту, и стоит
        # заметного времени. Считаем его только на полном синке: при scoped
        # (отчёт за период) и при инкременте — пропускаем (§5 спеки).
        if mode == "full":
            with profiler.stage("refresh_writeoff_stats"):
                project_card_service = ProjectCardService(request.bitrix24_account.client, request.bitrix24_account)
                project_card_service.refresh_writeoff_stats()
    except Exception:
        logger.exception("Timesheet sync failed for account %s", request.bitrix24_account.pk)
        profiler.set_metric("status", "error")
        profiler.log()
        return JsonResponse(
            {
                "status": "warning",
                "count": 0,
                "warning": "Не удалось обновить данные из Битрикс24. Используются последние сохраненные данные.",
            }
        )

    with profiler.stage("invalidate_caches"):
        invalidate_project_runtime_caches(request.bitrix24_account)
    request.bitrix24_account.last_timesheet_synced_at = now
    request.bitrix24_account.save(update_fields=["last_timesheet_synced_at"])
    profiler.set_metric("count", count)
    response = JsonResponse({"status": "success", "count": count, "last_synced_at": now.isoformat()})
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("timesheet_sync_status")
@auth_required
def timesheet_sync_status(request: AuthorizedRequest):
    acc = request.bitrix24_account
    count = TimesheetItem.objects.filter(**scope_to_tenant(acc)).count()
    last = acc.last_timesheet_synced_at
    return JsonResponse({
        "last_synced_at": last.isoformat() if last else None,
        "count": count,
    })


@xframe_options_exempt
@require_GET
@log_errors("timesheet_list")
@auth_required
def timesheet_list(request: AuthorizedRequest):
    queryset = TimesheetItem.objects.filter(**scope_to_tenant(request.bitrix24_account)).order_by('-created_at', '-bitrix_id')

    # Filter by record creation date (created_at)
    created_from = request.GET.get('created_from')
    created_to = request.GET.get('created_to')
    queryset = _apply_created_at_filters(queryset, created_from, created_to)

    page_number = request.GET.get('page', 1)
    page_size = _parse_page_size(request)

    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page_number)

    items = []
    for item in page_obj:
        items.append({
            "id": item.bitrix_id,
            "task_id": item.task_id,
            "employee_id": item.employee_id,
            "hours": item.hours,
            "is_billable": item.is_billable,
            "non_billable_hours": item.non_billable_hours,
            "description": item.description,
            "project_title": item.project_title,
            "date": item.date_reflection.isoformat() if item.date_reflection else None,
            "created_at": item.created_at.isoformat()
        })
        
    return JsonResponse({
        "items": items,
        "total": paginator.count,
        "page": page_obj.number,
        "pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
    })


@xframe_options_exempt
@require_GET
@log_errors("get_users")
@auth_required
def get_users(request: AuthorizedRequest):
    # order_by("last_name", "name") тайбрейкается "bitrix_id" (уникален в
    # пределах тенанта — unique_together на PortalUser): без него SQL не
    # гарантирует стабильный порядок между запросами разных страниц при
    # совпадении last_name/name, что даёт дубли/пропуски при постраничном
    # обходе с фронта (Задача 6). См. ревью Задачи 5, Important #2.
    queryset = PortalUser.objects.filter(**scope_to_tenant(request.bitrix24_account)).order_by("last_name", "name", "bitrix_id")

    active_only = str(request.GET.get("active_only", "")).strip().lower() in {"1", "true", "y", "yes"}
    if active_only:
        queryset = queryset.filter(active=True)

    page_number = request.GET.get("page", 1)
    # Клэмп сверху (200) не даёт ?limit=100000 сериализовать весь справочник
    # сотрудников в один ответ. См. ревью Задачи 5, Important #1.
    page_size = _parse_page_size(request)

    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page_number)

    items = [
        {
            "id": item.bitrix_id,
            "name": item.name,
            "last_name": item.last_name,
            "active": item.active,
            "updated_at": item.updated_at.isoformat(),
        }
        for item in page_obj
    ]

    return JsonResponse({
        "items": items,
        "total": paginator.count,
        "page": page_obj.number,
        "pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
    })


@xframe_options_exempt
@require_GET
@log_errors("get_configuration")
@auth_required
def get_configuration(request: AuthorizedRequest):
    service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse(service.get_configuration_sync())


@rate_limit("config_save_sync", 6, 60, key="account")
def _save_configuration_with_project_sync(
    request: AuthorizedRequest, service: ConfigurationService, config: dict
) -> JsonResponse:
    """Ветка save_configuration при заданном (> 0) project_sp_entity_type_id —

    вынесена отдельно, чтобы @rate_limit покрывал только её, тем же приёмом,
    что и _get_project_board_meta_refresh. Технически иначе: там флаг
    (?refresh=1) читается из query string и декоратор мог сработать сразу,
    до разбора тела. Здесь флаг зависит от тела POST-запроса, которое нужно
    сначала разобрать и нормализовать — этим занимается save_configuration
    (там же отрабатывает обработка кривого JSON/конфигурации, до всякого
    лимита), и только когда известно, что эта ветка будет выполняться,
    управление передаётся сюда.

    Внутри — _build_project_spa_validation_payload (несколько живых вызовов:
    crm.item.list x2, no-op crm.item.update для проверки прав на запись) и,
    если валидация не отклонила конфигурацию, ProjectSyncService.sync() — та
    же полная синхронизация, что и у sync_project_board, с безусловным живым
    crm.company.list внутри, — плюс backfill_timesheet_project_items().
    account_sync_lock здесь — Postgres advisory-lock (взаимное исключение),
    а не ограничитель: не даёт двум синкам идти параллельно, но никак не
    мешает слать сохранения подряд без остановки — захватил, отработал,
    отпустил, снова захватил. Поэтому нужен отдельный rate_limit.

    project_sp_entity_type_id приходит из тела запроса клиента и не является
    секретом (то же значение возвращает get_configuration) — значит любой
    запрос с валидным токеном может выставить его и звать эту ветку в цикле.

    Порог 6/60 — тот же класс риска и то же число, что у соседнего
    sync_project_board (@rate_limit("sync", 6, 60, key="account")): внутри
    вызывается тот же ProjectSyncService.sync(). Привязка Project SPA в
    настройках — операция первичной настройки, которую выполняют редко и
    осознанно (в отличие, например, от автокомплита company_search — 60/60):
    6 запросов в минуту с запасом покрывают ручной цикл «поправил
    маппинг -> сохранил -> проверил ошибку валидации -> сохранил снова».

    Отдельный scope ("config_save_sync", а не "sync") — общий бюджет с
    кнопкой «Синхронизировать» на доске проектов означал бы, что
    администратор, сохраняющий настройки при первичной привязке Project SPA,
    отбирает лимит у сотрудников, которые в этот момент работают с доской
    (и наоборот — серия ручных синков не должна мешать сохранить настройки).
    Разные сценарии с разной частотой — счётчики разные.

    Обычные сохранения без project_sp_entity_type_id в конфигурации вообще
    не доходят до этой функции (см. save_configuration) и не расходуют её
    бюджет — сколько угодно подряд.
    """
    warnings = []
    project_validation = None
    try:
        project_validation = _build_project_spa_validation_payload(service, request.bitrix24_account, config)
    except Exception as validation_exc:
        logger.exception("Configuration save validation failed: %s", validation_exc)
        warnings.append(
            "Проверка Project SPA временно недоступна. Настройки сохранены, "
            "но валидацию рекомендуется повторить позже."
        )

    if project_validation and not project_validation.get("is_valid"):
        return JsonResponse(
            {
                "status": "validation_error",
                "error": "Конфигурация Project SPA невалидна. Исправьте ошибки и повторите сохранение.",
                "validation": project_validation,
            },
            status=400,
        )

    service.save_configuration_sync(config)
    invalidate_project_runtime_caches(request.bitrix24_account)

    response_payload = {"status": "success"}
    project_sync_service = ProjectSyncService(request.bitrix24_account.client, request.bitrix24_account)
    try:
        with account_sync_lock(request.bitrix24_account, scope="project"):
            sync_result = project_sync_service.sync()
        response_payload["project_sync"] = sync_result
    except SyncLockBusy:
        warnings.append(
            "Синхронизация проектов уже выполняется, повторите позже."
        )
        response_payload["project_sync"] = {
            "status": "warning",
            "warning": "Синхронизация проектов уже выполняется, повторите позже.",
        }
    except Exception as sync_exc:
        logger.exception("Configuration save project sync failed: %s", sync_exc)
        warnings.append(
            "Настройки сохранены, но автосинхронизация проектов завершилась ошибкой."
        )
        response_payload["project_sync"] = {
            "status": "warning",
            "warning": "Автосинхронизация проектов завершилась ошибкой.",
        }

    try:
        backfill_result = project_sync_service.backfill_timesheet_project_items()
        response_payload["timesheet_backfill"] = backfill_result
    except Exception as backfill_exc:
        logger.exception("Configuration save timesheet backfill failed: %s", backfill_exc)
        warnings.append(
            "Настройки сохранены, но backfill связей меток времени завершился ошибкой."
        )
        response_payload["timesheet_backfill"] = {
            "status": "warning",
            "warning": "Backfill связей меток времени завершился ошибкой.",
        }

    if warnings:
        response_payload["warning"] = " ".join(warnings)

    return JsonResponse(response_payload)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("save_configuration")
@auth_required
def save_configuration(request: AuthorizedRequest):
    service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    try:
        body = json.loads(request.body)
        # json.loads может успешно разобрать НЕ-объект (список/число/строку/true/
        # null/[]) — тогда body.get(...) ниже упал бы AttributeError, пойманный
        # общим except Exception как 500, хотя виноват клиент. save_configuration
        # — операция с побочным эффектом (перезаписывает сохранённые настройки),
        # поэтому здесь честный 400, а не тихая деградация в {}.
        if not isinstance(body, dict):
            return JsonResponse({"error": "Некорректный формат тела запроса."}, status=400)
        config = body.get('config', {})
        if not isinstance(config, dict):
            return JsonResponse({"error": "Некорректный формат конфигурации."}, status=400)

        config = service.normalize_configuration_sync(config)

        try:
            should_validate_project_spa = int(config.get("project_sp_entity_type_id") or 0) > 0
        except (TypeError, ValueError):
            should_validate_project_spa = False

        if should_validate_project_spa:
            # project_sp_entity_type_id > 0 -> эта ветка попытается запустить
            # ProjectSyncService.sync() (полную синхронизацию с Битрикс) и
            # поэтому лимитируется отдельно — см. docstring
            # _save_configuration_with_project_sync. Ветка ниже (без Project
            # SPA в конфигурации) синк не запускает и не лимитируется вовсе.
            return _save_configuration_with_project_sync(request, service, config)

        service.save_configuration_sync(config)
        invalidate_project_runtime_caches(request.bitrix24_account)
        return JsonResponse({"status": "success"})
    except json.JSONDecodeError:
        return JsonResponse({"error": "Некорректное JSON тело запроса."}, status=400)
    except Exception:
        logger.exception("Configuration save failed")
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)


@xframe_options_exempt
@require_GET
@log_errors("get_internal_lists")
@auth_required
def get_internal_lists(request: AuthorizedRequest):
    service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    iblock_type_id = request.GET.get('iblockTypeId', 'lists')
    socnet_group_id = request.GET.get('socnetGroupId')
    socnet_group_value = int(socnet_group_id) if socnet_group_id and str(socnet_group_id).isdigit() else None
    return JsonResponse({"lists": service.get_internal_lists_sync(iblock_type_id=iblock_type_id, socnet_group_id=socnet_group_value)})


@xframe_options_exempt
@require_GET
@log_errors("get_smart_processes")
@auth_required
def get_smart_processes(request: AuthorizedRequest):
    service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    return JsonResponse({"types": service.get_smart_processes_sync()})


@xframe_options_exempt
@require_GET
@log_errors("get_sp_fields")
@auth_required
def get_sp_fields(request: AuthorizedRequest):
    service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    entity_type_id = request.GET.get('entityTypeId')
    if not entity_type_id:
         return JsonResponse({"fields": []})
    
    return JsonResponse({"fields": service.get_sp_fields_sync(int(entity_type_id))})


@xframe_options_exempt
@require_GET
@log_errors("get_project_spa_validation")
@auth_required
def get_project_spa_validation(request: AuthorizedRequest):
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()
    payload = _build_project_spa_validation_payload(config_service, request.bitrix24_account, config)
    return JsonResponse(payload)


@xframe_options_exempt
@require_GET
@log_errors("get_project_spa_stages")
@auth_required
def get_project_spa_stages(request: AuthorizedRequest):
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()
    entity_type_id_raw = request.GET.get("entityTypeId") or config.get("project_sp_entity_type_id") or 0
    try:
        entity_type_id = int(entity_type_id_raw)
    except (TypeError, ValueError):
        return JsonResponse({"error": "Некорректный entityTypeId"}, status=400)

    if not entity_type_id:
        return JsonResponse({"error": "Не выбран Смарт-процесс ПРОЕКТ"}, status=400)

    try:
        payload = config_service.get_project_spa_stages_sync(entity_type_id)
        return JsonResponse({"status": "success", **payload})
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("create_smart_process")
@auth_required
def create_smart_process(request: AuthorizedRequest):
    """Create a new Smart Process from settings page."""
    try:
        service = InstallationService(request.bitrix24_account.client, request.bitrix24_account)
        result = service.create_smart_process_only()
        return JsonResponse({"status": "success", **result})
    except InstallationError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except Exception:
        logger.exception("create operation failed")
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("create_fields")
@auth_required
def create_fields(request: AuthorizedRequest):
    """Create all required fields in the selected Smart Process."""
    import json as json_module
    try:
        body = json_module.loads(request.body)
        # json_module.loads может успешно разобрать НЕ-объект — body.get(...)
        # ниже упал бы AttributeError (пойман общим except Exception как 500).
        # entityTypeId и так обязателен и уже проверяется явно ниже — пустой
        # словарь просто заводит не-объектное тело в ту же ветку валидации.
        if not isinstance(body, dict):
            body = {}
        sp_id = body.get('entityTypeId')
        if not sp_id:
            return JsonResponse({"error": "Не указан ID смарт-процесса"}, status=400)

        service = InstallationService(request.bitrix24_account.client, request.bitrix24_account)
        result = service.create_fields_only(int(sp_id))
        return JsonResponse({"status": "success", **result})
    except InstallationError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except Exception:
        logger.exception("create operation failed")
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("create_mapped_field")
@auth_required
def create_mapped_field(request: AuthorizedRequest):
    import json as json_module
    try:
        body = json_module.loads(request.body)
        # См. create_fields выше — тот же приём: не-объектное тело -> {}, и
        # entityTypeId/fieldKey естественно попадают в уже существующую
        # проверку ниже вместо AttributeError из body.get(...).
        if not isinstance(body, dict):
            body = {}
        sp_id = body.get('entityTypeId')
        field_key = body.get('fieldKey')
        mapping_type = body.get('mappingType') or 'timesheet'

        if not sp_id:
            return JsonResponse({"error": "Не указан ID смарт-процесса"}, status=400)
        if not field_key:
            return JsonResponse({"error": "Не указан ключ поля"}, status=400)

        service = InstallationService(request.bitrix24_account.client, request.bitrix24_account)
        result = service.create_single_field(int(sp_id), str(field_key), str(mapping_type))
        return JsonResponse({"status": "success", **result})
    except InstallationError as e:
        # field_warnings — ответы Битрикса по каждому полю. На успешной ветке они
        # уже отдавались, на ветке отказа терялись, и 400 приходил без причины.
        payload = {"error": str(e)}
        warnings = getattr(e, "warnings", None)
        if warnings:
            payload["field_warnings"] = warnings
        return JsonResponse(payload, status=400)
    except Exception:
        logger.exception("create operation failed")
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)


@xframe_options_exempt
@require_GET
@log_errors("report_daily_workload")
@auth_required
def report_daily_workload(request: AuthorizedRequest):
    profiler = ReportProfiler("report_daily_workload", account_id=request.bitrix24_account.pk)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Defaults for date if missing (current month)
    from datetime import date
    if not date_from:
        today = date.today()
        date_from = date(today.year, today.month, 1).isoformat()
    if not date_to:
        today = date.today()
        date_to = today.isoformat()

    with profiler.stage("queryset_build"):
        queryset = build_filtered_timesheet_queryset(
            request.bitrix24_account,
            {
                "date_from": date_from,
                "date_to": date_to,
                "employee_ids[]": request.GET.getlist("employee_ids[]"),
                "project_ids[]": request.GET.getlist("project_ids[]"),
                "employee_mode": request.GET.get("employee_mode", "include"),
                "project_mode": request.GET.get("project_mode", "include"),
            },
        )
    with profiler.stage("project_lookup"):
        project_name_by_item, project_name_by_group = build_project_title_lookups(request.bitrix24_account)
    with profiler.stage("materialize"):
        rows = materialize_rows(
            queryset,
            (
                "employee_id",
                "project_item_id",
                "project_id",
                "project_title",
                "hours",
                "task_id",
                "task_hierarchy_titles",
                "description",
                "date_reflection",
            ),
        )
    user_ids = {row["employee_id"] for row in rows if row.get("employee_id")}
    with profiler.stage("user_map"):
        user_map = _get_user_map(request, user_ids)
    with profiler.stage("build_items"):
        items = [
            {
                "sotrudnik_id": row["employee_id"],
                "project_name": resolve_project_name_for_row(row, project_name_by_item, project_name_by_group),
                "kolichestvo_chasov": row["hours"],
                "id_zadachi": row["task_id"],
                "nazvanie_zadachi": row["task_hierarchy_titles"][-1] if row.get("task_hierarchy_titles") else "No Title",
                "opisanie": row["description"],
                "data": row["date_reflection"].isoformat() if row.get("date_reflection") else None,
            }
            for row in rows
        ]

    with profiler.stage("service_generate"):
        report_service = ReportService()
        report = report_service.generate_daily_workload(items, user_map, date_from, date_to)
    profiler.set_metric("rows", len(rows))
    profiler.set_metric("users", len(user_ids))
    with profiler.stage("serialize"):
        response = JsonResponse(report, safe=False)
    profiler.attach_to_response(response)
    profiler.log()
    return response


@xframe_options_exempt
@require_GET
@log_errors("get_request_logs")
@auth_required
def get_request_logs(request: AuthorizedRequest):
    page_number = request.GET.get('page', 1)
    # Диагностический админский эндпоинт: разбор инцидента выигрывает от
    # страницы больше, чем у пользовательских списков, а тело каждой записи
    # ограничено сверху (RequestLoggingMiddleware.MAX_BODY_LENGTH).
    page_size = _parse_page_size(request, max_value=LOG_PAGE_SIZE_MAX)

    queryset = RequestLog.objects.filter(bitrix24_account=request.bitrix24_account).order_by('-timestamp')
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page_number)
    
    items = []
    for item in page_obj:
        items.append({
            "id": str(item.id),
            "timestamp": item.timestamp.isoformat(),
            "method": item.method,
            "path": item.path,
            "status_code": item.status_code,
            "duration_ms": item.duration_ms,
            "request_body": str(item.request_body) if item.request_body else "",
            "response_body": str(item.response_body) if item.response_body else "",
        })
        
    return JsonResponse({
        "items": items,
        "total": paginator.count,
        "page": page_obj.number,
        "pages": paginator.num_pages,
    })


@xframe_options_exempt
@require_GET
@log_errors("get_system_logs")
@auth_required
def get_system_logs(request: AuthorizedRequest):
    page_number = request.GET.get('page', 1)
    page_size = _parse_page_size(request, max_value=LOG_PAGE_SIZE_MAX)

    queryset = SystemLog.objects.filter(bitrix24_account=request.bitrix24_account).order_by('-timestamp')
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page_number)
    
    items = []
    for item in page_obj:
        items.append({
            "id": str(item.id),
            "timestamp": item.timestamp.isoformat(),
            "level": item.level,
            "module": item.module,
            "message": item.message,
            "traceback": item.traceback or ""
        })
        
    return JsonResponse({
        "items": items,
        "total": paginator.count,
        "page": page_obj.number,
        "pages": paginator.num_pages,
    })


@xframe_options_exempt
@require_GET
@log_errors("inn_backfill_scan")
@auth_required
def inn_backfill_scan(request: AuthorizedRequest):
    """Поиск карточек списания без ИНН за период + предлагаемые значения из проекта."""
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()
    service = InnBackfillService(request.bitrix24_account.client, request.bitrix24_account, config)
    err = service.ensure_inn_fields()
    if err:
        return JsonResponse({"error": err}, status=400)
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    project_ids = request.GET.getlist("project_ids[]")
    result = service.scan(date_from, date_to, project_ids)
    return JsonResponse(result, safe=False)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("inn_backfill_apply")
@auth_required
def inn_backfill_apply(request: AuthorizedRequest):
    """Запись ИНН в поля OUR_INN/CLIENT_INN выбранных карточек списания (crm.item.update)."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError, UnicodeDecodeError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)
    # json.loads может успешно разобрать НЕ-объект — body.get(...) ниже упал бы
    # AttributeError необработанным (утечка сырого текста исключения клиенту
    # через log_errors). Запись ИНН необратима на стороне Bitrix — честный 400,
    # а не тихая деградация в {} (которая дала бы тот же 400, но по неверной
    # причине "карточки не выбраны").
    if not isinstance(body, dict):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)
    items = body.get("items", [])
    if not isinstance(items, list) or not items:
        return JsonResponse({"error": "Не переданы карточки для простановки"}, status=400)
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()
    service = InnBackfillService(request.bitrix24_account.client, request.bitrix24_account, config)
    err = service.ensure_inn_fields()
    if err:
        return JsonResponse({"error": err}, status=400)
    result = service.apply(items)
    return JsonResponse(result, safe=False)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("inn_backfill_project_items")
@auth_required
def inn_backfill_project_items(request: AuthorizedRequest):
    """Резолв карточек проекта для простановки/замены ИНН (запись делает фронт чанками)."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError, UnicodeDecodeError):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)
    # См. inn_backfill_apply выше: не-объектное тело проходит json.loads, но
    # body.get(...) ниже (после похода в ConfigurationService) упал бы
    # AttributeError необработанным. Тот же принцип — честный 400 сразу,
    # не дожидаясь похода в Bitrix за конфигурацией.
    if not isinstance(body, dict):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)
    config = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account).get_configuration_sync()
    service = InnBackfillService(request.bitrix24_account.client, request.bitrix24_account, config)
    err = service.ensure_inn_fields()
    if err:
        return JsonResponse({"error": err}, status=400)
    result = service.project_items(
        body.get("project_id", ""), body.get("date_from", ""), body.get("date_to", ""),
        body.get("our_inn", ""), body.get("client_inn", ""), bool(body.get("overwrite")),
    )
    return JsonResponse(result, safe=False)


@xframe_options_exempt
@require_GET
@log_errors("projects_health")
@auth_required
def projects_health(request: AuthorizedRequest):
    """Список незаполненных проектов (нет данных для ИНН)."""
    config = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account).get_configuration_sync()
    service = InnBackfillService(request.bitrix24_account.client, request.bitrix24_account, config)
    return JsonResponse(service.projects_health(), safe=False)


@xframe_options_exempt
@csrf_exempt
@require_POST
@log_errors("export_raw_data")
@auth_required
@rate_limit("export", 12, 60, key="account")
def export_raw_data(request: AuthorizedRequest):
    """
    Export raw CRM items from Bitrix24 Smart Process to Excel.
    Accepts JSON body with: date_from, date_to, date_type ('creation' or 'reflection'), fields (list of field IDs).
    - Employee fields are resolved to "Lastname Firstname" format.
    - All values are written as strings (number_format='@') to prevent type coercion.
    """
    try:
        body = json.loads(request.body)
    except Exception:
        # (json.JSONDecodeError, Exception) было избыточно: JSONDecodeError и
        # так подкласс Exception — упрощено при этой же правке.
        return JsonResponse({"error": "Invalid JSON body"}, status=400)
    # json.loads может успешно разобрать НЕ-объект — .get(...) на date_from/
    # date_type/fields ниже упал бы AttributeError необработанным. Это выгрузка
    # по явным фильтрам пользователя — честный 400, а не тихая выгрузка "всего
    # без фильтров" (что дал бы молчаливый {}).
    if not isinstance(body, dict):
        return JsonResponse({"error": "Invalid JSON body"}, status=400)

    date_from = body.get("date_from", "")
    date_to = body.get("date_to", "")
    date_type = body.get("date_type", "reflection")  # 'creation' or 'reflection'
    selected_fields = body.get("fields", [])

    # Load configuration to get entity_type_id and field mapping
    config_service = ConfigurationService(request.bitrix24_account.client, request.bitrix24_account)
    config = config_service.get_configuration_sync()

    entity_type_id = config.get("sp_entity_type_id", 0)
    if not entity_type_id:
        return JsonResponse({"error": "Смарт-процесс не настроен. Перейдите в Настройки и выберите смарт-процесс."}, status=400)

    # Get field definitions for header names and type info
    # Wrapped in try/except: if crm.item.fields fails it should not kill the whole export
    try:
        fields_meta = config_service.get_sp_fields_sync(int(entity_type_id))
    except Exception:
        logger.exception("export_raw_data: failed to fetch smart-process fields")
        return JsonResponse(
            {"error": "Внутренняя ошибка сервера"},
            status=500
        )

    field_labels = {f["id"]: f["title"] for f in fields_meta}

    # Identify employee-type fields so we can resolve IDs to names later
    employee_field_ids = {f["id"] for f in fields_meta if f.get("type") == "employee"}

    # If no specific fields requested, take all available
    if not selected_fields:
        selected_fields = [f["id"] for f in fields_meta]

    # Build Bitrix24 filter based on date_type
    crm_filter = _build_export_date_filter(
        date_type, date_from, date_to, config.get("fields_mapping", {})
    )

    # ---------------------------------------------------------------------------
    # Helpers — локальные, не зависят от Django/request
    # ---------------------------------------------------------------------------
    def _extract_raw_items(response: dict) -> list:
        """Извлекает список items из ответа crm.item.list (одиночного или батчевого)."""
        result = response.get("result", {})
        if isinstance(result, dict):
            items = result.get("items")
            if items is None:
                items = result.get("result", [])
        else:
            items = result
        return items if isinstance(items, list) else []

    def _raw_data_offset_fallback(bx_token, eid, f_filter, f_select, page_size=50) -> list:
        """Резервный offset-цикл (исходная реализация). Используется при сбое батч-выборки."""
        fallback_items = []
        fb_start = 0
        while True:
            fb_resp = bx_token.call_method(
                "crm.item.list",
                {
                    "entityTypeId": int(eid),
                    "filter": f_filter,
                    "select": f_select if f_select else ["*"],
                    "start": fb_start,
                },
            )
            fb_result = fb_resp.get("result", {})
            fb_page = fb_result.get("items", [])
            fallback_items.extend(fb_page)
            fb_total = fb_resp.get("total", fb_result.get("total", 0))
            fb_start += page_size
            if not fb_page or len(fb_page) < page_size or fb_start >= fb_total:
                break
        return fallback_items

    # ---------------------------------------------------------------------------
    # Fetch all items — батч-выборка по образцу _fetch_all_pages_batched
    # ---------------------------------------------------------------------------
    import logging as _logging
    _log = _logging.getLogger(__name__)

    bx_token = request.bitrix24_account.client._bitrix_token
    _select = selected_fields if selected_fields else ["*"]
    _eid = int(entity_type_id)

    all_items = []
    try:
        # Первая страница — получаем items + total
        first_response = bx_token.call_method(
            "crm.item.list",
            {
                "entityTypeId": _eid,
                "filter": crm_filter,
                "select": _select,
                "start": 0,
            },
        )
        first_items = _extract_raw_items(first_response)
        first_result = first_response.get("result", {})
        # total лежит на верхнем уровне ответа (BUG FIX #3)
        total = first_response.get(
            "total",
            first_result.get("total", 0) if isinstance(first_result, dict) else 0,
        )
        try:
            total = int(total)
        except (TypeError, ValueError):
            total = len(first_items)

        all_items = list(first_items)
        page_size = 50

        if total > page_size:
            # Строим батч для всех оставшихся офсетов одним HTTP-запросом
            offsets = list(range(page_size, total, page_size))
            methods = {
                f"p{off}": (
                    "crm.item.list",
                    {
                        "entityTypeId": _eid,
                        "filter": crm_filter,
                        "select": _select,
                        "start": off,
                    },
                )
                for off in offsets
            }
            _log.info(
                "export_raw_data: fetching %s batch-offset pages (total=%s)",
                len(offsets),
                total,
            )
            batches_resp = bx_token.call_batches(methods, halt=False)

            # batches_resp["result"]["result"][key] = содержимое response["result"]
            # одиночного crm.item.list, т.е. {"items": [...]}
            try:
                sub_results = batches_resp.get("result", {}).get("result", {})
            except Exception:
                sub_results = {}

            if not isinstance(sub_results, dict):
                try:
                    sub_results = {str(i): v for i, v in enumerate(sub_results)}
                except Exception:
                    sub_results = {}

            batch_items_count = 0
            for key, sub_result in sub_results.items():
                try:
                    page_items = _extract_raw_items({"result": sub_result})
                    all_items.extend(page_items)
                    batch_items_count += len(page_items)
                except Exception as exc:
                    _log.warning(
                        "export_raw_data: could not parse batch sub-result key=%s: %s",
                        key,
                        exc,
                    )

            # Оборонительная проверка: если батч вернул 0 при ненулевом total — fallback
            if batch_items_count == 0 and total > page_size:
                _log.warning(
                    "export_raw_data: batch returned 0 items (total=%s, offsets=%s); "
                    "falling back to offset pagination.",
                    total,
                    len(offsets),
                )
                all_items = _raw_data_offset_fallback(bx_token, _eid, crm_filter, _select)

    except Exception as e:
        # Батч-выборка упала полностью — пробуем резервный offset-цикл
        _log.warning(
            "export_raw_data: batched fetch failed (%s); falling back to offset pagination.",
            e,
        )
        try:
            all_items = _raw_data_offset_fallback(bx_token, _eid, crm_filter, _select)
        except Exception:
            logger.exception("export_raw_data: failed to fetch items from Bitrix24")
            return JsonResponse(
                {
                    "error": "Внутренняя ошибка сервера",
                },
                status=500,
            )

    # --- Resolve employee IDs to names ---
    # Collect unique user IDs only from employee fields that were selected
    selected_employee_fields = [fid for fid in selected_fields if fid in employee_field_ids]
    user_ids_to_fetch = set()
    for item in all_items:
        for fid in selected_employee_fields:
            val = item.get(fid)
            if isinstance(val, list):
                for v in val:
                    if v:
                        user_ids_to_fetch.add(str(v))
            elif val:
                user_ids_to_fetch.add(str(val))

    # Имена — из локального PortalUser тем же _get_user_map, что и 14 report_*
    # (Фаза 2 sync-offload). Раньше здесь был собственный батчинг user.get по 50
    # id: синхронные вызовы Bitrix на каждый экспорт, мимо scope_to_tenant и
    # нормализации id. Карта — {канонический id: "Фамилия Имя"}.
    user_map = _get_user_map(request, user_ids_to_fetch)

    def _employee_name(val) -> str:
        """Имя сотрудника или сырой id, если имени нет.

        _get_user_map ключей для нерезолвнутых id не отдаёт вовсе, поэтому
        фолбэк на сырой id держим здесь (в отличие от resolve_employee_name с
        его "Сотрудник <id>" — в выгрузке сырых данных нужен именно id).
        Ключи карты каноничные, поэтому сначала ищем по нормализованному id,
        затем по сырому значению.
        """
        raw = str(val)
        return user_map.get(extract_bitrix_user_id(val)) or user_map.get(raw) or raw

    def resolve_employee(val) -> str:
        """Convert employee field value to human-readable name string."""
        if isinstance(val, list):
            return "; ".join(_employee_name(v) for v in val if v is not None)
        elif val is not None:
            return _employee_name(val)
        return ""

    # --- Build Excel workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сырые данные"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    headers = [field_labels.get(fid, fid) for fid in selected_fields]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_safe_cell_text(str(header)))
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows — all values as strings, employee fields resolved to names
    for row_idx, item in enumerate(all_items, start=2):
        for col_idx, fid in enumerate(selected_fields, start=1):
            raw_value = item.get(fid, "")

            # Resolve employee fields to names
            if fid in employee_field_ids:
                str_value = resolve_employee(raw_value)
            elif isinstance(raw_value, (list, dict)):
                str_value = json.dumps(raw_value, ensure_ascii=False)
            elif raw_value is None:
                str_value = ""
            else:
                str_value = str(raw_value)

            cell = ws.cell(row=row_idx, column=col_idx, value=_safe_cell_text(str_value))
            # Force Excel to treat the cell as text (prevents large numbers → scientific notation)
            cell.number_format = '@'

    # Auto-adjust column widths
    for col in ws.columns:
        max_len: int = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(int(max_len) + 4, 50)

    # Return as HTTP response with Excel content type
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=raw_data_export.xlsx"
    return response


@xframe_options_exempt
@csrf_exempt
def serve_spa(request):
    """
    Serve index.html for any route (GET or POST) to support Bitrix24 iframe loading and SPA navigation.
    """
    logger.info("serve_spa called. Method=%s Path=%s", request.method, request.path)
    try:
        # settings.BASE_DIR points to /app inside container
        # index.html is copied to /app/frontend_build/index.html
        index_path = settings.BASE_DIR / "frontend_build" / "index.html"
        with open(index_path, 'r') as f:
            return HttpResponse(f.read())
    except FileNotFoundError:
        return HttpResponse(
            f"Frontend not found at {index_path}. "
            "Please ensure 'npm run generate' ran successfully during build.",
            status=404
        )

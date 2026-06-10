"""
Тесты безопасности: Excel-инъекция формул, CORS-regex, отключаемая Django-админка.
Задача 1.3 sprint-1-security.
"""
import io
import os
import re
import unittest

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "test_settings")
django.setup()

import openpyxl  # noqa: E402
from django.test import TestCase  # noqa: E402
from django.urls import resolve, Resolver404  # noqa: E402

from main.report_excel import (  # noqa: E402
    build_hierarchy_workbook,
    build_matrix_workbook,
    build_table_workbook,
    build_project_task_workbook,
)


# ---------------------------------------------------------------------------
# Вспомогательные символы-инъекции
# ---------------------------------------------------------------------------
_INJECTION_VALUES = ["=2+2", "+CMD()", "-1+1", "@SUM(A1)", "=HYPERLINK(\"http://evil.io\",\"click\")"]
_SAFE_VALUE = "Проект А"  # обычное значение — не должно меняться


# ---------------------------------------------------------------------------
# А) Excel-инъекция: build_project_task_workbook (через _write_row → node.name)
# ---------------------------------------------------------------------------
class TestExcelInjectionProjectTask(TestCase):
    def _get_cell_a4(self, workbook_bytes):
        wb = openpyxl.load_workbook(workbook_bytes)
        ws = wb.active
        # Строка 1 — title, 2 — subtitle, 3 — headers, 4 — первая строка данных
        return ws.cell(row=4, column=1)

    def test_injection_formula_gets_apostrophe_prefix(self):
        for payload in _INJECTION_VALUES:
            with self.subTest(payload=payload):
                nodes = [{"name": payload, "total_hours": 1, "billable_hours": 1, "non_billable_hours": 0}]
                out = build_project_task_workbook(nodes)
                cell = self._get_cell_a4(out)
                # Нейтрализованное значение начинается с апострофа ИЛИ data_type == 's' (строка)
                self.assertTrue(
                    str(cell.value).startswith("'"),
                    msg=f"Payload «{payload}» не нейтрализован: value={cell.value!r}"
                )

    def test_safe_name_unchanged(self):
        nodes = [{"name": _SAFE_VALUE, "total_hours": 1, "billable_hours": 1, "non_billable_hours": 0}]
        out = build_project_task_workbook(nodes)
        cell = self._get_cell_a4(out)
        self.assertEqual(cell.value, _SAFE_VALUE)


# ---------------------------------------------------------------------------
# А) Excel-инъекция: build_hierarchy_workbook (node name в _write)
# ---------------------------------------------------------------------------
class TestExcelInjectionHierarchy(TestCase):
    def _get_first_data_cell(self, out):
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Строка 1 — title, 2 — headers, 3 — первый узел
        return ws.cell(row=3, column=1)

    def test_injection_formula_gets_apostrophe_prefix(self):
        for payload in _INJECTION_VALUES:
            with self.subTest(payload=payload):
                roots = [{"name": payload, "total_hours": 1, "billable_hours": 1, "non_billable_hours": 0}]
                out = build_hierarchy_workbook(roots, title="Тест")
                cell = self._get_first_data_cell(out)
                self.assertTrue(
                    str(cell.value).startswith("'"),
                    msg=f"Payload «{payload}» не нейтрализован: value={cell.value!r}"
                )

    def test_safe_name_unchanged(self):
        roots = [{"name": _SAFE_VALUE, "total_hours": 1, "billable_hours": 1, "non_billable_hours": 0}]
        out = build_hierarchy_workbook(roots, title="Тест")
        cell = self._get_first_data_cell(out)
        self.assertEqual(cell.value, _SAFE_VALUE)


# ---------------------------------------------------------------------------
# А) Excel-инъекция: build_matrix_workbook (employee name)
# ---------------------------------------------------------------------------
class TestExcelInjectionMatrix(TestCase):
    def _get_employee_cell(self, out):
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Строка 1 — title, 2 — headers, 3 — первая строка данных
        return ws.cell(row=3, column=1)

    def test_injection_formula_gets_apostrophe_prefix(self):
        header_days = [{"date": "2026-05-01"}]
        for payload in _INJECTION_VALUES:
            with self.subTest(payload=payload):
                rows = [{"employee": {"name": payload}, "days": {"2026-05-01": {"total": 1}}}]
                out = build_matrix_workbook(header_days, rows, title="Тест")
                cell = self._get_employee_cell(out)
                self.assertTrue(
                    str(cell.value).startswith("'"),
                    msg=f"Payload «{payload}» не нейтрализован: value={cell.value!r}"
                )

    def test_safe_name_unchanged(self):
        header_days = [{"date": "2026-05-01"}]
        rows = [{"employee": {"name": _SAFE_VALUE}, "days": {"2026-05-01": {"total": 1}}}]
        out = build_matrix_workbook(header_days, rows, title="Тест")
        cell = self._get_employee_cell(out)
        self.assertEqual(cell.value, _SAFE_VALUE)


# ---------------------------------------------------------------------------
# А) Excel-инъекция: build_table_workbook (fmt="text" ветка в _put)
# ---------------------------------------------------------------------------
class TestExcelInjectionTable(TestCase):
    def _get_first_data_cell(self, out):
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Строка 1 — title, 2 — headers, 3 — первая строка данных
        return ws.cell(row=3, column=1)

    def test_injection_formula_gets_apostrophe_prefix(self):
        cols = [{"key": "name", "label": "Название", "fmt": "text"}]
        for payload in _INJECTION_VALUES:
            with self.subTest(payload=payload):
                rows = [{"name": payload}]
                out = build_table_workbook(cols, rows, title="Тест")
                cell = self._get_first_data_cell(out)
                self.assertTrue(
                    str(cell.value).startswith("'"),
                    msg=f"Payload «{payload}» не нейтрализован: value={cell.value!r}"
                )

    def test_safe_name_unchanged(self):
        cols = [{"key": "name", "label": "Название", "fmt": "text"}]
        rows = [{"name": _SAFE_VALUE}]
        out = build_table_workbook(cols, rows, title="Тест")
        cell = self._get_first_data_cell(out)
        self.assertEqual(cell.value, _SAFE_VALUE)


# ---------------------------------------------------------------------------
# Б) CORS-regex: тест константы BITRIX24_ORIGIN_REGEX из settings
# ---------------------------------------------------------------------------
class TestCorsRegex(TestCase):
    def setUp(self):
        from django.conf import settings
        # Импортируем напрямую из модуля settings (не через django.conf который может быть обёрткой)
        import settings as settings_module
        self.regex = getattr(settings_module, "BITRIX24_ORIGIN_REGEX", None)

    def test_bitrix24_origin_regex_constant_exists(self):
        self.assertIsNotNone(
            self.regex,
            "Константа BITRIX24_ORIGIN_REGEX должна быть определена в settings.py"
        )

    def _match(self, origin):
        return bool(re.match(self.regex, origin))

    def test_valid_origins_match(self):
        valid = [
            "https://acme.bitrix24.ru",
            "https://my-portal.bitrix24.com.br",
            "https://x.bitrix24.co.uk",
            "https://test.bitrix24.com",
            "https://company.bitrix24.de",
        ]
        for origin in valid:
            with self.subTest(origin=origin):
                self.assertTrue(self._match(origin), f"Должен матчиться: {origin}")

    def test_malicious_origins_do_not_match(self):
        invalid = [
            "https://a.bitrix24.com.attacker.io",   # поддомен после TLD
            "https://bitrix24.ru",                   # нет поддомена
            "http://acme.bitrix24.ru",               # HTTP вместо HTTPS
            "https://acme.bitrix24.evil",            # несуществующий TLD
            "https://bitrix24.com.attacker.io",      # вариант без субдомена b24
            "https://notbitrix24.ru",                # вообще не bitrix24
        ]
        for origin in invalid:
            with self.subTest(origin=origin):
                self.assertFalse(self._match(origin), f"Не должен матчиться: {origin}")


# ---------------------------------------------------------------------------
# В) Админка за флагом: при django_admin_enabled=False → /api/admin/ даёт 404
# ---------------------------------------------------------------------------
class TestAdminUrlDisabled(TestCase):
    def test_admin_url_not_reachable_when_disabled(self):
        """
        В тестовом окружении django_admin_enabled=False (default).
        Запрос GET /api/admin/ должен вернуть 404.
        """
        from config import config
        if config.django_admin_enabled:
            self.skipTest("Тест актуален только при django_admin_enabled=False")

        response = self.client.get("/api/admin/")
        self.assertEqual(
            response.status_code, 404,
            "URL /api/admin/ должен возвращать 404 когда DJANGO_ADMIN_ENABLED не установлен"
        )


if __name__ == "__main__":
    unittest.main()
